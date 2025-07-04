from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
import sqlite3
import pandas as pd
import random
import tempfile
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from typing import Optional

app = FastAPI(title="Attendance Report Generator", version="1.0.0")

@app.post("/generate-attendance-report")
async def generate_attendance_report(
    db_file: UploadFile = File(..., description="ZK.db SQLite database file"),
    start_date: str = Form(..., description="Start date in YYYY-MM format (e.g., 2025-06)"),
    end_date: Optional[str] = Form(None, description="End date in YYYY-MM format (optional, defaults to start_date)")
):
    """
    Generate Excel attendance report from ZK.db file for specified date range.
    
    Parameters:
    - db_file: ZK.db SQLite database file
    - start_date: Start date in YYYY-MM format
    - end_date: End date in YYYY-MM format (optional)
    """
    
    # Validate date format
    try:
        datetime.strptime(start_date, "%Y-%m")
        if end_date:
            datetime.strptime(end_date, "%Y-%m")
        else:
            end_date = start_date
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM format.")
    
    # Validate file type
    if not db_file.filename.endswith('.db'):
        raise HTTPException(status_code=400, detail="File must be a .db file")
    
    # Create temporary file for uploaded database
    with tempfile.NamedTemporaryFile(delete=False, suffix='.db') as temp_db:
        content = await db_file.read()
        temp_db.write(content)
        temp_db_path = temp_db.name
    
    try:
        # Connect to the temporary database
        conn = sqlite3.connect(temp_db_path)
        
        # Modified SQL query to use date parameters
        query = f"""
        WITH punches_per_day AS (
            SELECT 
                p.employee_id,
                date(p.punch_time) AS punch_date,
                time(p.punch_time) AS punch_time,
                p.punch_time AS full_punch_time
            FROM att_punches p
            WHERE strftime('%Y-%m', p.punch_time) >= '{start_date}'
            AND strftime('%Y-%m', p.punch_time) <= '{end_date}'
        ),

        ranked_punches AS (
            SELECT
                p.employee_id,
                p.punch_date,
                p.punch_time,
                ROW_NUMBER() OVER (PARTITION BY p.employee_id, p.punch_date ORDER BY p.full_punch_time ASC) AS rn
            FROM punches_per_day p
        ),

        final AS (
            SELECT
                e.emp_pin AS employee_id,
                e.emp_firstname || ' ' || COALESCE(e.emp_lastname, '') AS full_name,
                d.dept_name AS department,
                r.punch_date AS Date,
                CASE strftime('%w', r.punch_date)
                    WHEN '0' THEN 'Sun.'
                    WHEN '1' THEN 'Mon.'
                    WHEN '2' THEN 'Tues.'
                    WHEN '3' THEN 'Wed.'
                    WHEN '4' THEN 'Thur.'
                    WHEN '5' THEN 'Fri.'
                    WHEN '6' THEN 'Sat.'
                END AS Workday,
                tt.timetable_name AS Timetable,
                time(tt.timetable_start) AS StartWorkTime,
                time(tt.timetable_end) AS EndWorkTime,
                MAX(CASE WHEN r.rn = 1 THEN r.punch_time END) AS `Clock-In`,
                MAX(CASE WHEN r.rn = 2 THEN r.punch_time END) AS `Clock-Out`,
                MAX(CASE WHEN r.rn = 3 THEN r.punch_time END) AS `In`,
                MAX(CASE WHEN r.rn = 4 THEN r.punch_time END) AS `Out`
            FROM ranked_punches r
            JOIN hr_employee e ON e.id = r.employee_id
            LEFT JOIN hr_department d ON e.department_id = d.id
            LEFT JOIN att_day_details ad ON ad.employee_id = r.employee_id AND date(ad.att_date) = r.punch_date
            LEFT JOIN att_timetable tt ON ad.timetable_id = tt.id
            GROUP BY e.emp_pin, e.emp_firstname, e.emp_lastname, d.dept_name, r.punch_date, tt.timetable_name, tt.timetable_start, tt.timetable_end
        ),

        with_flags AS (
            SELECT 
                *,
                -- Late Clock In
                CASE 
                    WHEN time(`Clock-In`) > time(StartWorkTime)
                    THEN printf('%02d:%02d', 
                        (strftime('%s', time(`Clock-In`)) - strftime('%s', time(StartWorkTime))) / 3600,
                        ((strftime('%s', time(`Clock-In`)) - strftime('%s', time(StartWorkTime))) % 3600) / 60
                    )
                    ELSE '00:00'
                END AS `Late Clock In`,

                -- Early Clock In
                CASE 
                    WHEN time(`Clock-In`) < time(StartWorkTime)
                    THEN printf('%02d:%02d', 
                        (strftime('%s', time(StartWorkTime)) - strftime('%s', time(`Clock-In`))) / 3600,
                        ((strftime('%s', time(StartWorkTime)) - strftime('%s', time(`Clock-In`))) % 3600) / 60
                    )
                    ELSE '00:00'
                END AS `Early Clock In`,

                -- Early Clock Out: Use `Out` first, fallback to `Clock-Out`
                CASE 
                    WHEN (`Out` IS NOT NULL AND time(`Out`) < time(EndWorkTime))
                    THEN printf('%02d:%02d', 
                        (strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Out`))) / 3600,
                        ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Out`))) % 3600) / 60
                    )
                    WHEN (`Out` IS NULL AND `Clock-Out` IS NOT NULL AND time(`Clock-Out`) < time(EndWorkTime))
                    THEN printf('%02d:%02d', 
                        (strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Clock-Out`))) / 3600,
                        ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Clock-Out`))) % 3600) / 60
                    )
                    ELSE '00:00'
                END AS `Early Clock Out`,

                -- Break
                printf('%02d:%02d',
                    CASE
                        WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL AND (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`))) >= 0
                            THEN (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`)))
                        WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL
                            THEN (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`)) + 86400)
                        ELSE 0
                    END / 3600,
                    CASE
                        WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL AND (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`))) >= 0
                            THEN ((strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`))) % 3600) / 60
                        WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL
                            THEN ((strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`)) + 86400) % 3600) / 60
                        ELSE 0
                    END
                ) AS `Break`
            FROM final
        ),

        with_work_time AS (
            SELECT 
                *,
                -- Required Work Time
                printf('%02d:%02d',
                    CASE 
                        WHEN (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime))) >= 0 
                        THEN (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) - 3600)
                        ELSE (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) + 86400 - 3600)
                    END / 3600,
                    CASE 
                        WHEN (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime))) >= 0 
                        THEN ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) - 3600) % 3600) / 60
                        ELSE ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) + 86400 - 3600) % 3600) / 60
                    END
                ) AS `Required Work Time`,

                -- Work Time
                printf('%02d:%02d',
                    CASE
                        WHEN `Out` IS NOT NULL AND (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                            THEN (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) - 3600)
                        WHEN `Out` IS NOT NULL
                            THEN (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600)
                        WHEN `Clock-Out` IS NOT NULL AND (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                            THEN (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) - 3600)
                        WHEN `Clock-Out` IS NOT NULL
                            THEN (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600)
                        ELSE 0
                    END / 3600,
                    CASE
                        WHEN `Out` IS NOT NULL AND (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                            THEN ((strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) - 3600) % 3600) / 60
                        WHEN `Out` IS NOT NULL
                            THEN ((strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600) % 3600) / 60
                        WHEN `Clock-Out` IS NOT NULL AND (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                            THEN ((strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) - 3600) % 3600) / 60
                        WHEN `Clock-Out` IS NOT NULL
                            THEN ((strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600) % 3600) / 60
                        ELSE 0
                    END
                ) AS `Work Time`
            FROM with_flags
        ),

        final_with_ot AS (
            SELECT 
                *,
                CASE 
                    WHEN `Clock-In` IS NOT NULL OR `Clock-Out` IS NOT NULL
                    THEN '00:00'
                    ELSE `Required Work Time`
                END AS `Absent`,

                CASE 
                    WHEN Workday IN ('Mon.', 'Tues.', 'Wed.', 'Thur.', 'Fri.')
                         AND (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) > 0
                    THEN printf('%02d:%02d',
                        (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) / 3600,
                        ((strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) % 3600) / 60
                    )
                    ELSE '00:00'
                END AS `OT1`,

                CASE 
                    WHEN Workday IN ('Sat.', 'Sun.')
                         AND (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) > 0
                    THEN printf('%02d:%02d',
                        (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) / 3600,
                        ((strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) % 3600) / 60
                    )
                    ELSE '00:00'
                END AS `OT2`,

                '00:00' AS `OT3`
            FROM with_work_time
        )
        SELECT 
            employee_id,
            full_name,
            department,
            Date,
            Workday,
            Timetable,
            `Required Work Time`,
            StartWorkTime,
            EndWorkTime,
            `Clock-In`,
            `Clock-Out`,
            `In`,
            `Out`,
            `Late Clock In`,
            `Early Clock In`,
            `Early Clock Out`,
            `Break`,
            `Work Time`,
            `Absent`,
            `OT1`,
            `OT2`,
            `OT3`
        FROM final_with_ot
        ORDER BY employee_id, Date;
        """
        
        df = pd.read_sql_query(query, conn)
        
        if df.empty:
            raise HTTPException(status_code=404, detail="No attendance data found for the specified date range")
        
        # Get company name for title
        title_query = "SELECT cmp_name FROM hr_company LIMIT 1;"
        title_df = pd.read_sql_query(title_query, conn)
        company_name = title_df.iloc[0]['cmp_name'] if not title_df.empty else "Company Name"
        
        conn.close()
        
        # Generate Excel file using the same logic as saya.py
        output_file = generate_excel_report(df, company_name, start_date, end_date)
        
        return FileResponse(
            path=output_file,
            filename=f"attendance_report_{start_date}_to_{end_date}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except sqlite3.Error as e:
        raise HTTPException(status_code=400, detail=f"Database error: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    finally:
        # Clean up temporary database file
        if os.path.exists(temp_db_path):
            os.unlink(temp_db_path)

def generate_excel_report(df, company_name, start_date, end_date):
    """Generate Excel report using the same formatting logic as saya.py"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Create fonts and fills (same as saya.py)
    tahoma_font = Font(name='Tahoma', size=10)
    tahoma_bold_font = Font(name='Tahoma', size=10, bold=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
    bright_red_font = Font(name='Tahoma', size=10, color='FF0000')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    header_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    total_base_fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
    total_day_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    clock_header_fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
    ot_header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    penalty_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    title_font = Font(name='Tahoma', size=22, bold=True)
    subtitle_font = Font(name='Tahoma', size=14, bold=True)

    current_row = 1

    # Add company name title at the top
    ws.cell(row=current_row, column=1, value=company_name).font = title_font
    current_row += 2

    # Add subtitle with date range
    ws.cell(row=current_row, column=1, value=f"Monthly Statement Report ({start_date} to {end_date})").font = subtitle_font
    current_row += 2

    # Write column headers
    cols_to_show = ['Date', 'Workday', 'Timetable', 'EYEE NAME', 'StartWorkTime', 'EndWorkTime', 'Clock-In', 'Clock-Out', 'In', 'Out', 'Required Work Time', 'Break', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Work Time', 'Absent', 'Penalty', 'OT1', 'OT2', 'OT3', 'Night Shift', 'Allowence', 'Total Base', 'Day', 'H', 'MC', 'AL', 'UP', 'S', 'Total Day']
    colored_header_columns = ['Date', 'Workday', 'Timetable', 'EYEE NAME', 'StartWorkTime', 'EndWorkTime', 'Day', 'H', 'MC', 'AL', 'UP', 'S', 'Required Work Time', 'Break', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Work Time', 'Absent']
    
    for col_idx, col_name in enumerate(cols_to_show, start=1):
        header_cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        header_cell.font = tahoma_bold_font
        header_cell.alignment = Alignment(wrap_text=True)
        
        # Apply background color to specific columns
        if col_name in colored_header_columns:
            header_cell.fill = header_fill
        elif col_name == 'Total Base':
            header_cell.fill = total_base_fill
        elif col_name in ['Total Day', 'Night Shift', 'Allowence']:
            header_cell.fill = total_day_fill
        elif col_name in ['Clock-In', 'Clock-Out', 'In', 'Out']:
            header_cell.fill = clock_header_fill
        elif col_name in ['OT1', 'OT2', 'OT3']:
            header_cell.fill = ot_header_fill
        elif col_name == 'Penalty':
            header_cell.fill = penalty_fill
    
    ws.row_dimensions[current_row].height = 39.75
    current_row += 1

    # Function to format time from HH:MM:SS to HH:MM
    def format_time(time_str):
        if not time_str or time_str == '':
            return time_str
        try:
            if time_str.count(':') == 2:
                parts = time_str.split(':')
                return f"{parts[0]}:{parts[1]}"
            else:
                return time_str
        except:
            return time_str

    # Process each employee
    for emp_id, group in df.groupby('employee_id'):
        emp_info = group.iloc[0]

        # Write employee info
        ws.cell(row=current_row, column=1, value="Employee ID").font = tahoma_bold_font
        ws.cell(row=current_row, column=2, value=emp_id).font = tahoma_font
        ws.cell(row=current_row, column=3, value="Full Name").font = tahoma_bold_font
        ws.cell(row=current_row, column=4, value=emp_info['full_name']).font = tahoma_font
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Write data rows (simplified version of the original logic)
        for _, row in group.iterrows():
            is_sunday = row.get('Workday') == 'Sun.'
            
            for col_idx, col_name in enumerate(cols_to_show, start=1):
                # Set value based on column name
                if col_name == 'EYEE NAME':
                    cell_value = emp_info['full_name']
                elif col_name == 'Penalty':
                    cell_value = '0.0'
                elif col_name == 'Night Shift':
                    original_timetable = row.get('Timetable', "")
                    cell_value = '2.0' if 'NIGHT' in str(original_timetable).upper() else '0.0'
                elif col_name == 'Allowence':
                    cell_value = '0.0'
                elif col_name == 'Total Base':
                    cell_value = '0.0' if is_sunday else '1.0'
                elif col_name == 'Day':
                    if is_sunday:
                        cell_value = ''
                    else:
                        clock_in = row.get('Clock-In', '')
                        clock_out = row.get('Clock-Out', '')
                        cell_value = '1.0' if (clock_in or clock_out) else ''
                elif col_name in ['H', 'MC', 'AL', 'UP', 'S']:
                    cell_value = ''
                elif col_name == 'Total Day':
                    cell_value = '1.0'
                elif col_name in ['StartWorkTime', 'EndWorkTime', 'Clock-In', 'Clock-Out', 'In', 'Out']:
                    cell_value = format_time(row.get(col_name, ""))
                else:
                    cell_value = row.get(col_name, "")
                
                cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                cell.font = tahoma_font
                
                # Apply basic coloring (simplified)
                if col_name == 'Penalty' and cell_value and cell_value != '' and not is_sunday:
                    cell.fill = penalty_fill
                elif col_name == 'Total Base' and cell_value and cell_value != '' and not is_sunday:
                    cell.fill = total_base_fill
                elif col_name in ['Total Day', 'Night Shift', 'Allowence'] and cell_value and cell_value != '' and not is_sunday:
                    cell.fill = total_day_fill
                elif is_sunday:
                    cell.fill = yellow_fill
            
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 2  # Add space after each employee

    # Apply borders
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(cols_to_show)):
        for cell in row:
            cell.border = thin_border

    # Save file
    random_num = random.randint(1000, 9999)
    filename = f"attendance_report_{random_num}.xlsx"
    wb.save(filename)
    
    return filename

@app.get("/")
async def root():
    """Root endpoint with API information"""
    return {
        "message": "Attendance Report Generator API",
        "version": "1.0.0",
        "endpoints": {
            "POST /generate-attendance-report": "Generate Excel attendance report from ZK.db file"
        }
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 