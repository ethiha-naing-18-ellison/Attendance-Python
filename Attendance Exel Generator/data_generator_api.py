from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
import sqlite3
import pandas as pd
import random
import tempfile
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from typing import Optional

app = FastAPI(title="Attendance Data Sheet Generator", version="1.0.0")

def filter_last_two_punches(df):
    """
    Check ALL consecutive punch pairs in each row.
    Different logic for first half vs second half:
    
    FIRST HALF (positions 0-3): Keep FIRST punch (earlier time)
    - Punch 1 → Punch 2: If < 10 min, keep Punch 1, remove Punch 2
    - Punch 2 → Punch 3: If < 10 min, keep Punch 2, remove Punch 3
    - Punch 3 → Punch 4: If < 10 min, keep Punch 3, remove Punch 4
    
    SECOND HALF (positions 4-5): Keep SECOND punch (later time)
    - Punch 4 → Punch 5: If < 10 min, keep Punch 5, remove Punch 4
    - Punch 5 → Punch 6: If < 10 min, keep Punch 6, remove Punch 5
    
    Example:
    Input:  08:30, 08:32, 12:00, 20:06, 20:07
            ^^^^^^^^^^^^^ Keep 08:30 (first)
                                ^^^^^^^^^^^^^ Keep 20:07 (second)
    Output: 08:30, 12:00, 20:07
    """
    def time_to_seconds(time_str):
        """Convert HH:MM time to seconds"""
        if not time_str or pd.isna(time_str) or time_str == '':
            return None
        try:
            parts = str(time_str).split(':')
            hours = int(parts[0])
            minutes = int(parts[1])
            return hours * 3600 + minutes * 60
        except:
            return None
    
    punch_cols = ['punch_1', 'punch_2', 'punch_3', 'punch_4', 'punch_5', 'punch_6']
    
    # Process each row
    for idx in df.index:
        # Keep checking until no more duplicates found
        changed = True
        while changed:
            changed = False
            
            # Collect all non-empty punches
            punches = []
            for col in punch_cols:
                punch_value = df.loc[idx, col]
                if punch_value and not pd.isna(punch_value) and str(punch_value).strip() != '':
                    punches.append(punch_value)
            
            # Check each consecutive pair
            for i in range(len(punches) - 1):
                current_time = punches[i]
                next_time = punches[i + 1]
                
                # Convert to seconds
                current_seconds = time_to_seconds(current_time)
                next_seconds = time_to_seconds(next_time)
                
                # Check if both are valid and less than 10 minutes (600 seconds) apart
                if current_seconds is not None and next_seconds is not None:
                    time_diff = abs(next_seconds - current_seconds)
                    
                    # If less than 10 minutes apart, it's a duplicate
                    if time_diff < 600:
                        # Decide which punch to remove based on position
                        # First half (positions 0-3): Keep first punch (remove second)
                        # Second half (positions 4-5): Keep second punch (remove first)
                        
                        if i < 3:  # First half: positions 0→1, 1→2, 2→3
                            # Keep current (first), remove next (second)
                            punches.pop(i + 1)
                        else:  # Second half: positions 3→4, 4→5
                            # Keep next (second), remove current (first)
                            punches.pop(i)
                        
                        changed = True
                        break  # Start checking again from the beginning
            
            # Rebuild the punch sequence without duplicates
            # Clear all punch columns first
            for col in punch_cols:
                df.loc[idx, col] = None
            
            # Reassign cleaned punches
            for i, punch in enumerate(punches):
                if i < len(punch_cols):
                    df.loc[idx, punch_cols[i]] = punch
    
    return df

@app.post("/generate-data-sheet")
async def generate_data_sheet(
    db_file: UploadFile = File(..., description="ZK.db SQLite database file"),
    start_date: str = Form(..., description="Start date in YYYY-MM-DD format (e.g., 2025-01-01)"),
    end_date: Optional[str] = Form(None, description="End date in YYYY-MM-DD format (optional, defaults to start_date)")
):
    """
    Generate Excel data sheet with raw punch times from ZK.db file.
    
    Columns: Employee ID | Name | In | Out | In | Out | In | Out
    """
    
    # Validate date format
    try:
        datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            datetime.strptime(end_date, "%Y-%m-%d")
        else:
            end_date = start_date
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD format.")
    
    # Validate file type
    if not db_file.filename.endswith('.db'):
        raise HTTPException(status_code=400, detail="File must be a .db file")
    
    # Create temporary file for uploaded database
    with tempfile.NamedTemporaryFile(delete=False, suffix='.db') as temp_db:
        content = await db_file.read()
        temp_db.write(content)
        temp_db_path = temp_db.name
    
    try:
        # Connect to the temporary database
        conn = sqlite3.connect(temp_db_path)
        
        # Get raw punch data with 10-minute rule (filters duplicate punches)
        raw_punch_query = f"""
        WITH punches_per_day AS (
            SELECT 
                p.employee_id,
                date(p.punch_time) AS punch_date,
                time(p.punch_time) AS punch_time,
                p.punch_time AS full_punch_time
            FROM att_punches p
            WHERE date(p.punch_time) >= '{start_date}'
            AND date(p.punch_time) <= '{end_date}'
        ),
        
        ranked_punches AS (
            SELECT
                p.employee_id,
                p.punch_date,
                p.punch_time,
                p.full_punch_time,
                ROW_NUMBER() OVER (PARTITION BY p.employee_id, p.punch_date ORDER BY p.full_punch_time ASC) AS rn
            FROM punches_per_day p
        ),
        
        initial_punches AS (
            SELECT 
                employee_id,
                punch_date,
                MAX(CASE WHEN rn = 1 THEN punch_time END) AS punch_1,
                MAX(CASE WHEN rn = 2 THEN punch_time END) AS punch_2,
                MAX(CASE WHEN rn = 3 THEN punch_time END) AS punch_3,
                MAX(CASE WHEN rn = 4 THEN punch_time END) AS punch_4,
                MAX(CASE WHEN rn = 5 THEN punch_time END) AS punch_5,
                MAX(CASE WHEN rn = 6 THEN punch_time END) AS punch_6,
                MAX(CASE WHEN rn = 7 THEN punch_time END) AS punch_7,
                MAX(CASE WHEN rn = 1 THEN full_punch_time END) AS full_punch_1,
                MAX(CASE WHEN rn = 2 THEN full_punch_time END) AS full_punch_2
            FROM ranked_punches
            GROUP BY employee_id, punch_date
        ),
        
        filtered_punches AS (
            SELECT 
                employee_id,
                punch_date,
                punch_1,
                -- Apply 10-minute rule: if punch_2 is less than 10 minutes after punch_1, skip it
                CASE 
                    WHEN full_punch_1 IS NOT NULL AND full_punch_2 IS NOT NULL 
                         AND (strftime('%s', full_punch_2) - strftime('%s', full_punch_1)) < 600
                    THEN punch_3  -- Skip punch_2, use punch_3 instead
                    ELSE punch_2  -- Use punch_2 normally
                END AS punch_2,
                CASE 
                    WHEN full_punch_1 IS NOT NULL AND full_punch_2 IS NOT NULL 
                         AND (strftime('%s', full_punch_2) - strftime('%s', full_punch_1)) < 600
                    THEN punch_4  -- Shift punch_4 to position 3
                    ELSE punch_3  -- Use punch_3 normally
                END AS punch_3,
                CASE 
                    WHEN full_punch_1 IS NOT NULL AND full_punch_2 IS NOT NULL 
                         AND (strftime('%s', full_punch_2) - strftime('%s', full_punch_1)) < 600
                    THEN punch_5  -- Shift punch_5 to position 4
                    ELSE punch_4  -- Use punch_4 normally
                END AS punch_4,
                CASE 
                    WHEN full_punch_1 IS NOT NULL AND full_punch_2 IS NOT NULL 
                         AND (strftime('%s', full_punch_2) - strftime('%s', full_punch_1)) < 600
                    THEN punch_6  -- Shift punch_6 to position 5
                    ELSE punch_5  -- Use punch_5 normally
                END AS punch_5,
                CASE 
                    WHEN full_punch_1 IS NOT NULL AND full_punch_2 IS NOT NULL 
                         AND (strftime('%s', full_punch_2) - strftime('%s', full_punch_1)) < 600
                    THEN punch_7  -- Shift punch_7 to position 6
                    ELSE punch_6  -- Use punch_6 normally
                END AS punch_6
            FROM initial_punches
        )
        
        SELECT 
            e.emp_pin AS employee_id,
            e.emp_firstname || ' ' || COALESCE(e.emp_lastname, '') AS full_name,
            fp.punch_date AS Date,
            fp.punch_1,
            fp.punch_2,
            fp.punch_3,
            fp.punch_4,
            fp.punch_5,
            fp.punch_6
        FROM filtered_punches fp
        JOIN hr_employee e ON e.emp_pin = (
            SELECT emp_pin FROM hr_employee WHERE id = fp.employee_id LIMIT 1
        )
        ORDER BY employee_id, Date;
        """
        
        raw_punch_df = pd.read_sql_query(raw_punch_query, conn)
        
        if raw_punch_df.empty:
            raise HTTPException(status_code=404, detail="No punch data found for the specified date range")
        
        # Get company name
        title_query = "SELECT cmp_name FROM hr_company LIMIT 1;"
        title_df = pd.read_sql_query(title_query, conn)
        company_name = title_df.iloc[0]['cmp_name'] if not title_df.empty else "Company Name"
        
        conn.close()
        
        # Apply 10-minute rule to LAST two punches (check for duplicate at end)
        raw_punch_df = filter_last_two_punches(raw_punch_df)
        
        # Generate Excel file
        output_file = create_data_excel(raw_punch_df, company_name, start_date, end_date)
        
        return FileResponse(
            path=output_file,
            filename=f"attendance_data_{start_date}_to_{end_date}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except sqlite3.Error as e:
        raise HTTPException(status_code=400, detail=f"Database error: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    finally:
        # Clean up temporary database file
        if os.path.exists(temp_db_path):
            os.unlink(temp_db_path)

def create_data_excel(raw_punch_df, company_name, start_date, end_date):
    """Generate Excel file with Data sheet only"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Styling
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    tahoma_font = Font(name='Tahoma', size=10)
    tahoma_bold_font = Font(name='Tahoma', size=10, bold=True)
    title_font = Font(name='Tahoma', size=22, bold=True)
    subtitle_font = Font(name='Tahoma', size=14, bold=True)
    date_range_font = Font(name='Tahoma', size=12, bold=True)
    header_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Blue fill for employee separator rows
    blue_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')  # Light blue
    
    current_row = 1
    
    # Add company name title
    ws.cell(row=current_row, column=1, value=company_name).font = title_font
    current_row += 2
    
    # Add subtitle
    ws.cell(row=current_row, column=1, value="Raw Punch Data Report").font = subtitle_font
    current_row += 1
    
    # Add date range
    date_range_text = f"{start_date} to {end_date}"
    ws.cell(row=current_row, column=1, value=date_range_text).font = date_range_font
    current_row += 2
    
    # Column headers
    headers = ['Employee ID', 'Name', 'In', 'Out', 'In', 'Out', 'In', 'Out']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = tahoma_bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal='center')
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # Function to format time from HH:MM:SS to HH:MM
    def format_time(time_str):
        if not time_str or time_str == '' or pd.isna(time_str):
            return ''
        try:
            if time_str.count(':') == 2:
                parts = time_str.split(':')
                return f"{parts[0]}:{parts[1]}"
            else:
                return time_str
        except:
            return ''
    
    # Write data rows
    if not raw_punch_df.empty:
        current_employee = None
        for _, row in raw_punch_df.iterrows():
            # Check if this is a new employee
            if current_employee is not None and current_employee != row.get('employee_id', ''):
                # Add blue separator row between employees
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col_idx, value='')
                    cell.fill = blue_fill
                    cell.border = thin_border
                ws.row_dimensions[current_row].height = 5  # Thin separator
                current_row += 1
            
            # Update current employee
            current_employee = row.get('employee_id', '')
            # Check if Sunday for yellow highlighting
            date_str = str(row.get('Date', ''))
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                is_sunday = date_obj.weekday() == 6  # 6 = Sunday
            except:
                is_sunday = False
            
            # Employee ID
            cell = ws.cell(row=current_row, column=1, value=row.get('employee_id', ''))
            cell.font = tahoma_font
            cell.border = thin_border
            if is_sunday:
                cell.fill = yellow_fill
            
            # Employee Name
            cell = ws.cell(row=current_row, column=2, value=row.get('full_name', ''))
            cell.font = tahoma_font
            cell.border = thin_border
            if is_sunday:
                cell.fill = yellow_fill
            
            # Punch times (In, Out, In, Out, In, Out) - using punch_1 through punch_6
            punch_columns = ['punch_1', 'punch_2', 'punch_3', 'punch_4', 'punch_5', 'punch_6']
            for col_idx, punch_col in enumerate(punch_columns, start=3):
                punch_value = format_time(row.get(punch_col, ''))
                cell = ws.cell(row=current_row, column=col_idx, value=punch_value)
                cell.font = tahoma_font
                cell.border = thin_border
                if is_sunday:
                    cell.fill = yellow_fill
            
            ws.row_dimensions[current_row].height = 18
            current_row += 1
    
    # Freeze panes at header row
    ws.freeze_panes = 'A7'
    
    # Set column widths
    ws.column_dimensions['A'].width = 12  # Employee ID
    ws.column_dimensions['B'].width = 25  # Name
    ws.column_dimensions['C'].width = 10  # In
    ws.column_dimensions['D'].width = 10  # Out
    ws.column_dimensions['E'].width = 10  # In
    ws.column_dimensions['F'].width = 10  # Out
    ws.column_dimensions['G'].width = 10  # In
    ws.column_dimensions['H'].width = 10  # Out
    
    # Save Excel file
    random_num = random.randint(1000, 9999)
    filename = f"attendance_data_{random_num}.xlsx"
    wb.save(filename)
    
    return filename

@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve the HTML upload form"""
    try:
        with open("upload_form.html", "r", encoding="utf-8") as f:
            html_content = f.read()
        return HTMLResponse(content=html_content)
    except FileNotFoundError:
        return HTMLResponse(content="""
        <html>
            <head>
                <title>Attendance Data Sheet Generator</title>
                <style>
                    body { font-family: Arial, sans-serif; max-width: 800px; margin: 50px auto; padding: 20px; }
                    h1 { color: #333; }
                    .info { background: #f0f0f0; padding: 15px; border-radius: 5px; margin: 20px 0; }
                </style>
            </head>
            <body>
                <h1>📊 Attendance Data Sheet Generator API</h1>
                <div class="info">
                    <p><strong>Upload form not found.</strong></p>
                    <p>Please use the API endpoints directly:</p>
                    <ul>
                        <li><strong>POST /generate-data-sheet</strong>: Generate Excel data sheet with raw punch times</li>
                        <li><strong>GET /docs</strong>: Interactive API documentation</li>
                    </ul>
                </div>
                <p><a href="/docs">→ Open Interactive API Documentation</a></p>
            </body>
        </html>
        """)

@app.get("/api")
async def api_info():
    """API information endpoint"""
    return {
        "message": "Attendance Data Sheet Generator API",
        "version": "1.0.0",
        "description": "Generates Excel files with raw punch data only",
        "columns": ["Employee ID", "Name", "In", "Out", "In", "Out", "In", "Out"],
        "endpoints": {
            "GET /": "Web interface for uploading files",
            "POST /generate-data-sheet": "Generate Excel data sheet from ZK.db file",
            "GET /docs": "Interactive API documentation"
        }
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)

