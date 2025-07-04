import sqlite3
import pandas as pd
import random
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
# Connect to your SQLite DB
conn = sqlite3.connect('ZK.db')  # Change this to your DB path

# Your long SQL query
query = """
WITH punches_per_day AS (
    SELECT 
        p.employee_id,
        date(p.punch_time) AS punch_date,
        time(p.punch_time) AS punch_time,
        p.punch_time AS full_punch_time
    FROM att_punches p
    WHERE strftime('%Y-%m', p.punch_time) = '2025-06'
),

ranked_punches AS (
    SELECT
        p.employee_id,
        p.punch_date,
        p.punch_time,
        ROW_NUMBER() OVER (PARTITION BY p.employee_id, p.punch_date ORDER BY p.full_punch_time ASC) AS rn
    FROM punches_per_day p
),

final AS (
    SELECT
        e.emp_pin AS employee_id,
        e.emp_firstname || ' ' || COALESCE(e.emp_lastname, '') AS full_name,
        d.dept_name AS department,
        r.punch_date AS Date,
        CASE strftime('%w', r.punch_date)
            WHEN '0' THEN 'Sun.'
            WHEN '1' THEN 'Mon.'
            WHEN '2' THEN 'Tues.'
            WHEN '3' THEN 'Wed.'
            WHEN '4' THEN 'Thur.'
            WHEN '5' THEN 'Fri.'
            WHEN '6' THEN 'Sat.'
        END AS Workday,
        tt.timetable_name AS Timetable,
        time(tt.timetable_start) AS StartWorkTime,
        time(tt.timetable_end) AS EndWorkTime,
        MAX(CASE WHEN r.rn = 1 THEN r.punch_time END) AS `Clock-In`,
        MAX(CASE WHEN r.rn = 2 THEN r.punch_time END) AS `Clock-Out`,
        MAX(CASE WHEN r.rn = 3 THEN r.punch_time END) AS `In`,
        MAX(CASE WHEN r.rn = 4 THEN r.punch_time END) AS `Out`
    FROM ranked_punches r
    JOIN hr_employee e ON e.id = r.employee_id
    LEFT JOIN hr_department d ON e.department_id = d.id
    LEFT JOIN att_day_details ad ON ad.employee_id = r.employee_id AND date(ad.att_date) = r.punch_date
    LEFT JOIN att_timetable tt ON ad.timetable_id = tt.id
    GROUP BY e.emp_pin, e.emp_firstname, e.emp_lastname, d.dept_name, r.punch_date, tt.timetable_name, tt.timetable_start, tt.timetable_end
),

with_flags AS (
    SELECT 
        *,
        -- Late Clock In
        CASE 
            WHEN time(`Clock-In`) > time(StartWorkTime)
            THEN printf('%02d:%02d', 
                (strftime('%s', time(`Clock-In`)) - strftime('%s', time(StartWorkTime))) / 3600,
                ((strftime('%s', time(`Clock-In`)) - strftime('%s', time(StartWorkTime))) % 3600) / 60
            )
            ELSE '00:00'
        END AS `Late Clock In`,

        -- Early Clock In
        CASE 
            WHEN time(`Clock-In`) < time(StartWorkTime)
            THEN printf('%02d:%02d', 
                (strftime('%s', time(StartWorkTime)) - strftime('%s', time(`Clock-In`))) / 3600,
                ((strftime('%s', time(StartWorkTime)) - strftime('%s', time(`Clock-In`))) % 3600) / 60
            )
            ELSE '00:00'
        END AS `Early Clock In`,

        -- Early Clock Out: Use `Out` first, fallback to `Clock-Out`
        CASE 
            WHEN (`Out` IS NOT NULL AND time(`Out`) < time(EndWorkTime))
            THEN printf('%02d:%02d', 
                (strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Out`))) / 3600,
                ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Out`))) % 3600) / 60
            )
            WHEN (`Out` IS NULL AND `Clock-Out` IS NOT NULL AND time(`Clock-Out`) < time(EndWorkTime))
            THEN printf('%02d:%02d', 
                (strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Clock-Out`))) / 3600,
                ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Clock-Out`))) % 3600) / 60
            )
            ELSE '00:00'
        END AS `Early Clock Out`,

        -- Break
        printf('%02d:%02d',
            CASE
                WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL AND (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`))) >= 0
                    THEN (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`)))
                WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL
                    THEN (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`)) + 86400)
                ELSE 0
            END / 3600,
            CASE
                WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL AND (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`))) >= 0
                    THEN ((strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`))) % 3600) / 60
                WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL
                    THEN ((strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`)) + 86400) % 3600) / 60
                ELSE 0
            END
        ) AS `Break`
    FROM final
),

with_work_time AS (
    SELECT 
        *,
        -- Required Work Time
        printf('%02d:%02d',
            CASE 
                WHEN (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime))) >= 0 
                THEN (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) - 3600)
                ELSE (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) + 86400 - 3600)
            END / 3600,
            CASE 
                WHEN (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime))) >= 0 
                THEN ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) - 3600) % 3600) / 60
                ELSE ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) + 86400 - 3600) % 3600) / 60
            END
        ) AS `Required Work Time`,

        -- Work Time
        printf('%02d:%02d',
            CASE
                WHEN `Out` IS NOT NULL AND (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                    THEN (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) - 3600)
                WHEN `Out` IS NOT NULL
                    THEN (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600)
                WHEN `Clock-Out` IS NOT NULL AND (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                    THEN (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) - 3600)
                WHEN `Clock-Out` IS NOT NULL
                    THEN (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600)
                ELSE 0
            END / 3600,
            CASE
                WHEN `Out` IS NOT NULL AND (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                    THEN ((strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) - 3600) % 3600) / 60
                WHEN `Out` IS NOT NULL
                    THEN ((strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600) % 3600) / 60
                WHEN `Clock-Out` IS NOT NULL AND (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                    THEN ((strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) - 3600) % 3600) / 60
                WHEN `Clock-Out` IS NOT NULL
                    THEN ((strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600) % 3600) / 60
                ELSE 0
            END
        ) AS `Work Time`
    FROM with_flags
),

final_with_ot AS (
    SELECT 
        *,
        CASE 
            WHEN `Clock-In` IS NOT NULL OR `Clock-Out` IS NOT NULL
            THEN '00:00'
            ELSE `Required Work Time`
        END AS `Absent`,

        CASE 
            WHEN Workday IN ('Mon.', 'Tues.', 'Wed.', 'Thur.', 'Fri.')
                 AND (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) > 0
            THEN printf('%02d:%02d',
                (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) / 3600,
                ((strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) % 3600) / 60
            )
            ELSE '00:00'
        END AS `OT1`,

        CASE 
            WHEN Workday IN ('Sat.', 'Sun.')
                 AND (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) > 0
            THEN printf('%02d:%02d',
                (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) / 3600,
                ((strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) % 3600) / 60
            )
            ELSE '00:00'
        END AS `OT2`,

        '00:00' AS `OT3`
    FROM with_work_time
)
SELECT 
    employee_id,
    full_name,
    department,
    Date,
    Workday,
    Timetable,
    `Required Work Time`,
    StartWorkTime,
    EndWorkTime,
    `Clock-In`,
    `Clock-Out`,
    `In`,
    `Out`,
    `Late Clock In`,
    `Early Clock In`,
    `Early Clock Out`,
    `Break`,
    `Work Time`,
    `Absent`,
    `OT1`,
    `OT2`,
    `OT3`
FROM final_with_ot
ORDER BY employee_id, Date;

"""
df = pd.read_sql_query(query, conn)

# Get company name for title
title_query = "SELECT cmp_name FROM hr_company LIMIT 1;"
title_df = pd.read_sql_query(title_query, conn)
company_name = title_df.iloc[0]['cmp_name'] if not title_df.empty else "Company Name"

conn.close()

# Columns to sum
sum_columns = ['Required Work Time', 'Work Time', 'Absent', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Penalty', 'OT1', 'OT2', 'OT3', 'Night Shift', 'Allowence', 'Total Base', 'Day', 'H', 'MC', 'AL', 'UP', 'S', 'Total Day']

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Attendance"

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Create Tahoma font
tahoma_font = Font(name='Tahoma', size=10)
tahoma_bold_font = Font(name='Tahoma', size=10, bold=True)

# Create yellow fill for Sunday rows
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Create light red fill for late clock-in cells
red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')

# Create light purple fill for penalty cells
light_purple_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

# Create light orange fill for total base cells
light_orange_fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')

# Create bright red font for suspicious early clock in
bright_red_font = Font(name='Tahoma', size=10, color='FF0000')

# Create orange fill for suspicious punch pattern rows
orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

# Create green fill for total rows
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

# Create light yellow fill for specific header columns
header_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

# Create light orange fill for Total Base column
total_base_fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')

# Create light green fill for Total Day column
total_day_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

# Create light orange fill for Clock columns
clock_header_fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')

# Create light blue fill for OT columns
ot_header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

# Create light purple fill for Penalty column
penalty_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

# Create title font
title_font = Font(name='Tahoma', size=22, bold=True)

# Create subtitle font
subtitle_font = Font(name='Tahoma', size=14, bold=True)

current_row = 1

# Add company name title at the top
ws.cell(row=current_row, column=1, value=company_name).font = title_font
current_row += 2  # Add 1 empty row of space after title

# Add subtitle
ws.cell(row=current_row, column=1, value="Monthly Statement Report").font = subtitle_font
current_row += 2  # Add 1 empty row of space after subtitle

# Write column headers only once
cols_to_show = ['Date', 'Workday', 'Timetable', 'EYEE NAME', 'StartWorkTime', 'EndWorkTime', 'Clock-In', 'Clock-Out', 'In', 'Out', 'Required Work Time', 'Break', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Work Time', 'Absent', 'Penalty', 'OT1', 'OT2', 'OT3', 'Night Shift', 'Allowence', 'Total Base', 'Day', 'H', 'MC', 'AL', 'UP', 'S', 'Total Day']
# Define columns that need the special header color
colored_header_columns = ['Date', 'Workday', 'Timetable', 'EYEE NAME', 'StartWorkTime', 'EndWorkTime', 'Day', 'H', 'MC', 'AL', 'UP', 'S', 'Required Work Time', 'Break', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Work Time', 'Absent']
for col_idx, col_name in enumerate(cols_to_show, start=1):
    header_cell = ws.cell(row=current_row, column=col_idx, value=col_name)
    header_cell.font = tahoma_bold_font
    header_cell.alignment = Alignment(wrap_text=True)
    # Apply background color to specific columns
    if col_name in colored_header_columns:
        header_cell.fill = header_fill
    elif col_name == 'Total Base':
        header_cell.fill = total_base_fill
    elif col_name in ['Total Day', 'Night Shift', 'Allowence']:
        header_cell.fill = total_day_fill
    elif col_name in ['Clock-In', 'Clock-Out', 'In', 'Out']:
        header_cell.fill = clock_header_fill
    elif col_name in ['OT1', 'OT2', 'OT3']:
        header_cell.fill = ot_header_fill
    elif col_name == 'Penalty':
        header_cell.fill = penalty_fill
ws.row_dimensions[current_row].height = 39.75
current_row += 1

for emp_id, group in df.groupby('employee_id'):
    emp_info = group.iloc[0]

    # Write employee info in single row format
    ws.cell(row=current_row, column=1, value="Employee ID").font = tahoma_bold_font
    ws.cell(row=current_row, column=2, value=emp_id).font = tahoma_font
    ws.cell(row=current_row, column=3, value="Full Name").font = tahoma_bold_font
    ws.cell(row=current_row, column=4, value=emp_info['full_name']).font = tahoma_font
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    # Function to format time from HH:MM:SS to HH:MM
    def format_time(time_str):
        if not time_str or time_str == '':
            return time_str
        try:
            # If time has seconds, remove them
            if time_str.count(':') == 2:
                parts = time_str.split(':')
                return f"{parts[0]}:{parts[1]}"
            else:
                return time_str
        except:
            return time_str

    # Write data rows
    data_start_row = current_row  # Mark the start of data rows for grouping
    for _, row in group.iterrows():
        is_sunday = row.get('Workday') == 'Sun.'
        
        # Function to convert time string to minutes for comparison
        def time_to_minutes(time_str):
            if not time_str or time_str in ['', '0:00', '00:00']:
                return 0
            try:
                parts = time_str.split(':')
                hours = int(parts[0])
                minutes = int(parts[1])
                return hours * 60 + minutes
            except:
                return 0
        
        # Check if this is a suspicious row (Early Clock In > 2:30)
        early_in_value = row.get('Early Clock In', "")
        early_in_minutes = time_to_minutes(early_in_value)
        is_suspicious_row = early_in_minutes > 150  # 2:30 = 150 minutes
        
        # Check for suspicious punch patterns
        def has_value(val):
            return val and val.strip() != ''
        
        clock_in = row.get('Clock-In', '')
        clock_out = row.get('Clock-Out', '')
        in_punch = row.get('In', '')
        out_punch = row.get('Out', '')
        
        clock_in_exists = has_value(clock_in)
        clock_out_exists = has_value(clock_out)
        in_exists = has_value(in_punch)
        out_exists = has_value(out_punch)
        
        # Determine if punch pattern is suspicious
        is_punch_suspicious = False
        if clock_in_exists and not clock_out_exists and not in_exists and not out_exists:
            # Only Clock In exists = SUSPICIOUS
            is_punch_suspicious = True
        elif clock_in_exists and clock_out_exists and in_exists and not out_exists:
            # Clock In + Clock Out + In (missing Out) = SUSPICIOUS
            is_punch_suspicious = True
        
        for col_idx, col_name in enumerate(cols_to_show, start=1):
            # Set value based on column name
            if col_name == 'EYEE NAME':
                cell_value = emp_info['full_name']
            elif col_name == 'Timetable':
                # Format as "Timetable (StartTime - EndTime)"
                timetable_name = row.get('Timetable', "")
                start_time = format_time(row.get('StartWorkTime', ""))
                end_time = format_time(row.get('EndWorkTime', ""))
                if timetable_name and start_time and end_time:
                    cell_value = f"{timetable_name} ({start_time} - {end_time})"
                else:
                    cell_value = timetable_name
            elif col_name == 'Penalty':
                cell_value = '0.0'
            elif col_name == 'Night Shift':
                # Check if original Timetable contains "NIGHT"
                original_timetable = row.get('Timetable', "")
                if 'NIGHT' in str(original_timetable).upper():
                    cell_value = '2.0'
                else:
                    cell_value = '0.0'
            elif col_name == 'Allowence':
                cell_value = '0.0'
            elif col_name == 'Total Base':
                # Set 1.0 for non-Sunday, 0.0 for Sunday
                if is_sunday:
                    cell_value = '0.0'
                else:
                    cell_value = '1.0'
            elif col_name == 'Day':
                # Sunday is always empty, other days depend on clock-in/out presence
                if is_sunday:
                    cell_value = ''
                else:
                    # Check if worker was present (has clock-in OR clock-out)
                    clock_in = row.get('Clock-In', '')
                    clock_out = row.get('Clock-Out', '')
                    if clock_in or clock_out:  # If either clock-in or clock-out has value
                        cell_value = '1.0'
                    else:
                        cell_value = ''
            elif col_name in ['H', 'MC', 'AL', 'UP', 'S']:
                # Leave these columns empty
                cell_value = ''
            elif col_name == 'Total Day':
                # Fill every cell with 1.0
                cell_value = '1.0'
            elif col_name in ['StartWorkTime', 'EndWorkTime', 'Clock-In', 'Clock-Out', 'In', 'Out']:
                # Format time columns to show only HH:MM (remove seconds)
                cell_value = format_time(row.get(col_name, ""))
            else:
                cell_value = row.get(col_name, "")
            
            # Convert OT time values to decimal format
            if col_name in ['OT1', 'OT2', 'OT3'] and cell_value:
                # Convert "hh:mm" to decimal (e.g., "02:30" -> "2.5")
                def time_to_decimal(time_str):
                    if not time_str or time_str in ['', '0:00', '00:00']:
                        return '0.0'
                    try:
                        parts = time_str.split(':')
                        hours = int(parts[0])
                        minutes = int(parts[1])
                        decimal_value = hours + (minutes / 60.0)
                        return f"{decimal_value:.2f}".rstrip('0').rstrip('.')
                    except:
                        return cell_value
                
                cell_value = time_to_decimal(cell_value)
            
            cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
            
            # Apply font styling based on suspicious row detection
            if is_suspicious_row and col_name in ['Timetable', 'StartWorkTime', 'EndWorkTime', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Night Shift']:
                cell.font = bright_red_font  # Apply bright red font for suspicious rows
            else:
                cell.font = tahoma_font  # Apply regular font to data cells
            
            # Apply cell coloring based on conditions (punch suspicious takes priority)
            if is_punch_suspicious:
                # Apply orange background to entire row for suspicious punch patterns
                cell.fill = orange_fill
            elif col_name == 'Late Clock In':
                # Check if this is a late clock-in cell with a value other than 0:00
                late_value = row.get(col_name, "")
                if late_value and late_value not in ['0:00', '00:00', '']:
                    cell.fill = red_fill
                elif is_sunday:
                    cell.fill = yellow_fill
            elif col_name == 'Early Clock In':
                # Apply Sunday background if applicable
                if is_sunday:
                    cell.fill = yellow_fill
            elif col_name == 'Early Clock Out':
                # Check if this is an early clock-out cell with a value other than 0:00
                early_out_value = row.get(col_name, "")
                if early_out_value and early_out_value not in ['0:00', '00:00', '']:
                    cell.fill = red_fill  # Light red even on Sunday
                elif is_sunday:
                    cell.fill = yellow_fill
            elif col_name == 'Penalty':
                # Check if this is a penalty cell
                if is_sunday:
                    cell.fill = yellow_fill
                elif cell_value and cell_value != '':
                    # Apply light pink color to all non-empty, non-Sunday cells
                    cell.fill = penalty_fill
            elif col_name == 'Total Base':
                # Check if this is a total base cell
                if is_sunday:
                    cell.fill = yellow_fill
                elif cell_value and cell_value != '':
                    # Apply new orange color to all non-empty, non-Sunday cells
                    cell.fill = total_base_fill
            elif col_name == 'Total Day':
                # Check if this is a total day cell
                if is_sunday:
                    cell.fill = yellow_fill
                elif cell_value and cell_value != '':
                    # Apply light green color to all non-empty, non-Sunday cells
                    cell.fill = total_day_fill
            elif col_name in ['Night Shift', 'Allowence']:
                # Check if this is a night shift or allowence cell
                if is_sunday:
                    cell.fill = yellow_fill
                elif cell_value and cell_value != '':
                    # Apply light green color to all non-empty, non-Sunday cells
                    cell.fill = total_day_fill
            elif is_sunday:
                cell.fill = yellow_fill
        ws.row_dimensions[current_row].height = 18
        # Set outline level for data rows (level 1 for grouping)
        ws.row_dimensions[current_row].outline_level = 1
        current_row += 1

    # Write total row
    total_cell = ws.cell(row=current_row, column=1, value="TOTAL")
    total_cell.font = tahoma_bold_font
    total_cell.fill = green_fill
    for col_idx, col_name in enumerate(cols_to_show, start=1):
        if col_name in sum_columns:
            if col_name in ['Penalty', 'Night Shift', 'Allowence', 'Total Base', 'Day', 'H', 'MC', 'AL', 'UP', 'S', 'Total Day', 'OT1', 'OT2', 'OT3']:
                # For decimal columns, sum the decimal values
                if col_name == 'Penalty':
                    # For Penalty column, since all values are 0.0, total is 0.0
                    penalty_cell = ws.cell(row=current_row, column=col_idx, value="0.0")
                    penalty_cell.font = tahoma_font
                    penalty_cell.fill = green_fill
                elif col_name == 'Night Shift':
                    # Sum the night shift values (0.0 or 2.0)
                    total_night_shift = 0.0
                    for _, data_row in group.iterrows():
                        original_timetable = data_row.get('Timetable', "")
                        if 'NIGHT' in str(original_timetable).upper():
                            total_night_shift += 2.0
                    night_shift_cell = ws.cell(row=current_row, column=col_idx, value=f"{total_night_shift:.1f}")
                    night_shift_cell.font = tahoma_font
                    night_shift_cell.fill = green_fill
                elif col_name == 'Allowence':
                    # For Allowence column, since all values are 0.0, total is 0.0
                    allowence_cell = ws.cell(row=current_row, column=col_idx, value="0.0")
                    allowence_cell.font = tahoma_font
                    allowence_cell.fill = green_fill
                elif col_name == 'Total Base':
                    # For Total Base column, sum 1.0 for non-Sunday and 0.0 for Sunday
                    total_base = 0.0
                    for _, data_row in group.iterrows():
                        workday_value = data_row.get('Workday', "")
                        if workday_value != 'Sun.':
                            total_base += 1.0
                        # Sunday rows contribute 0.0 (no need to add)
                    total_base_cell = ws.cell(row=current_row, column=col_idx, value=f"{total_base:.1f}")
                    total_base_cell.font = tahoma_font
                    total_base_cell.fill = green_fill
                elif col_name == 'Day':
                    # For Day column, sum 1.0 for non-Sunday days where worker was present
                    total_days = 0.0
                    for _, data_row in group.iterrows():
                        workday_value = data_row.get('Workday', "")
                        if workday_value != 'Sun.':  # Not Sunday
                            # Check if worker was present (has clock-in OR clock-out)
                            clock_in = data_row.get('Clock-In', '')
                            clock_out = data_row.get('Clock-Out', '')
                            if clock_in or clock_out:  # If either clock-in or clock-out has value
                                total_days += 1.0
                        # Sunday rows contribute 0.0 (no need to add)
                    day_cell = ws.cell(row=current_row, column=col_idx, value=f"{total_days:.1f}")
                    day_cell.font = tahoma_font
                    day_cell.fill = green_fill
                elif col_name in ['H', 'MC', 'AL', 'UP', 'S']:
                    # For these columns, since all values are empty, total is empty
                    empty_cell = ws.cell(row=current_row, column=col_idx, value="")
                    empty_cell.font = tahoma_font
                    empty_cell.fill = green_fill
                elif col_name == 'Total Day':
                    # For Total Day column, sum all 1.0 values
                    total_day_count = len(group)  # Count all rows (each has 1.0)
                    total_day_cell = ws.cell(row=current_row, column=col_idx, value=f"{total_day_count:.1f}")
                    total_day_cell.font = tahoma_font
                    total_day_cell.fill = green_fill
                elif col_name in ['OT1', 'OT2', 'OT3']:
                    # For OT columns, sum the decimal values (converted from time)
                    total_ot = 0.0
                    for _, data_row in group.iterrows():
                        ot_value = data_row.get(col_name, "")
                        if ot_value and ot_value not in ['', '0:00', '00:00']:
                            # Convert time to decimal for summing
                            try:
                                parts = ot_value.split(':')
                                hours = int(parts[0])
                                minutes = int(parts[1])
                                decimal_value = hours + (minutes / 60.0)
                                total_ot += decimal_value
                            except:
                                pass
                    ot_cell = ws.cell(row=current_row, column=col_idx, value=f"{total_ot:.2f}".rstrip('0').rstrip('.'))
                    ot_cell.font = tahoma_font
                    ot_cell.fill = green_fill
            else:
                # Convert "hh:mm" to minutes for summing
                def time_to_minutes(t):
                    h, m = map(int, t.split(":"))
                    return h * 60 + m

                total_minutes = sum(time_to_minutes(v) for v in group[col_name] if isinstance(v, str) and ":" in v)
                total_h = total_minutes // 60
                total_m = total_minutes % 60
                time_cell = ws.cell(row=current_row, column=col_idx, value=f"{total_h}:{total_m:02}")
                time_cell.font = tahoma_font
                time_cell.fill = green_fill
        elif col_idx > 1:  # Don't override the TOTAL label in column 1
            # Apply regular font to empty cells in total row (except column 1)
            empty_total_cell = ws.cell(row=current_row, column=col_idx, value="")
            empty_total_cell.font = tahoma_font
            empty_total_cell.fill = green_fill
    ws.row_dimensions[current_row].height = 18

    # Add one empty row after each employee
    current_row += 2

# Apply border to all cells with data (excluding title rows)
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(cols_to_show)):
    for cell in row:
        cell.border = thin_border

# Freeze panes to keep header rows visible when scrolling
# Freeze at the row right after the column headers (row 6)
ws.freeze_panes = 'A6'

# Set outline groups to be collapsed by default
# Hide all data rows (outline level 1) by default to show collapsed view
for row_num in range(1, ws.max_row + 1):
    if row_num in ws.row_dimensions:
        if hasattr(ws.row_dimensions[row_num], 'outline_level') and ws.row_dimensions[row_num].outline_level == 1:
            ws.row_dimensions[row_num].hidden = True

# Save Excel file with random number
random_num = random.randint(1000, 9999)
filename = f"employee_attendance_grouped_{random_num}.xlsx"
wb.save(filename)
print(f"✅ Excel file created: {filename}")