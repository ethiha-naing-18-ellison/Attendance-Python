from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
import sqlite3
import pandas as pd
import random
import tempfile
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from typing import Optional

app = FastAPI(title="Attendance Report Generator", version="1.0.0")

@app.post("/generate-attendance-report")
async def generate_attendance_report(
    db_file: UploadFile = File(..., description="ZK.db SQLite database file"),
    start_date: str = Form(..., description="Start date in YYYY-MM-DD format (e.g., 2025-06-01)"),
    end_date: Optional[str] = Form(None, description="End date in YYYY-MM-DD format (optional, defaults to start_date)")
):
    """
    Generate Excel attendance report from ZK.db file for specified date range.
    
    Parameters:
    - db_file: ZK.db SQLite database file
    - start_date: Start date in YYYY-MM-DD format
    - end_date: End date in YYYY-MM-DD format (optional)
    """
    
    # Validate date format
    try:
        datetime.strptime(start_date, "%Y-%m-%d")
        if end_date:
            datetime.strptime(end_date, "%Y-%m-%d")
        # If end_date is not provided, we'll set it to the latest date in database later
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD format.")
    
    # Validate file type
    if not db_file.filename.endswith('.db'):
        raise HTTPException(status_code=400, detail="File must be a .db file")
    
    # Create temporary file for uploaded database
    with tempfile.NamedTemporaryFile(delete=False, suffix='.db') as temp_db:
        content = await db_file.read()
        temp_db.write(content)
        temp_db_path = temp_db.name
    
    try:
        # Connect to the temporary database
        conn = sqlite3.connect(temp_db_path)
        
        # If end_date is not provided, find the latest date in the database
        if not end_date:
            max_date_query = """
            SELECT MAX(date(punch_time)) as max_punch_date
            FROM att_punches
            UNION
            SELECT MAX(date(att_date)) as max_att_date
            FROM att_day_details
            ORDER BY max_punch_date DESC
            LIMIT 1
            """
            max_date_result = pd.read_sql_query(max_date_query, conn)
            if not max_date_result.empty and max_date_result.iloc[0]['max_punch_date']:
                end_date = max_date_result.iloc[0]['max_punch_date']
                print(f"DEBUG: end_date was empty, set to latest database date: {end_date}")
            else:
                # Fallback to start_date if no data found
                end_date = start_date
                print(f"DEBUG: No data found in database, using start_date as end_date: {end_date}")
        
        # Modified SQL query to use date parameters and include all dates in range
        query = f"""
        WITH RECURSIVE date_range AS (
            SELECT '{start_date}' AS date_value
            UNION ALL
            SELECT date(date_value, '+1 day')
            FROM date_range
            WHERE date_value < '{end_date}'
        ),

        employees_in_period AS (
            SELECT DISTINCT e.id as employee_id, e.emp_pin, e.emp_firstname, e.emp_lastname, e.department_id
            FROM hr_employee e
            WHERE e.id IN (
                SELECT DISTINCT employee_id 
                FROM att_punches 
                WHERE date(punch_time) >= '{start_date}' AND date(punch_time) <= '{end_date}'
                UNION
                SELECT DISTINCT employee_id 
                FROM att_day_details 
                WHERE date(att_date) >= '{start_date}' AND date(att_date) <= '{end_date}'
            )
        ),

        all_employee_dates AS (
            SELECT 
                ep.employee_id,
                ep.emp_pin,
                ep.emp_firstname, 
                ep.emp_lastname,
                ep.department_id,
                dr.date_value as punch_date
            FROM date_range dr
            CROSS JOIN employees_in_period ep
        ),

        punches_per_day AS (
            SELECT 
                p.employee_id,
                date(p.punch_time) AS punch_date,
                time(p.punch_time) AS punch_time,
                p.punch_time AS full_punch_time
            FROM att_punches p
            WHERE date(p.punch_time) >= '{start_date}'
            AND date(p.punch_time) <= '{end_date}'
        ),

        ranked_punches AS (
            SELECT
                p.employee_id,
                p.punch_date,
                p.punch_time,
                ROW_NUMBER() OVER (PARTITION BY p.employee_id, p.punch_date ORDER BY p.full_punch_time ASC) AS rn
            FROM punches_per_day p
        ),

        final AS (
            SELECT
                aed.emp_pin AS employee_id,
                aed.emp_firstname || ' ' || COALESCE(aed.emp_lastname, '') AS full_name,
                d.dept_name AS department,
                aed.punch_date AS Date,
                CASE strftime('%w', aed.punch_date)
                    WHEN '0' THEN 'Sun.'
                    WHEN '1' THEN 'Mon.'
                    WHEN '2' THEN 'Tues.'
                    WHEN '3' THEN 'Wed.'
                    WHEN '4' THEN 'Thur.'
                    WHEN '5' THEN 'Fri.'
                    WHEN '6' THEN 'Sat.'
                END AS Workday,
                tt.timetable_name AS Timetable,
                time(tt.timetable_start) AS StartWorkTime,
                time(tt.timetable_end) AS EndWorkTime,
                MAX(CASE WHEN r.rn = 1 THEN r.punch_time END) AS `Clock-In`,
                MAX(CASE WHEN r.rn = 2 THEN r.punch_time END) AS `Clock-Out`,
                MAX(CASE WHEN r.rn = 3 THEN r.punch_time END) AS `In`,
                MAX(CASE WHEN r.rn = 4 THEN r.punch_time END) AS `Out`
            FROM all_employee_dates aed
            LEFT JOIN ranked_punches r ON r.employee_id = aed.employee_id AND r.punch_date = aed.punch_date
            LEFT JOIN hr_department d ON aed.department_id = d.id
            LEFT JOIN att_day_details ad ON ad.employee_id = aed.employee_id AND date(ad.att_date) = aed.punch_date
            LEFT JOIN att_timetable tt ON ad.timetable_id = tt.id
            GROUP BY aed.emp_pin, aed.emp_firstname, aed.emp_lastname, d.dept_name, aed.punch_date, tt.timetable_name, tt.timetable_start, tt.timetable_end
        ),

        with_flags AS (
            SELECT 
                *,
                -- Late Clock In
                CASE 
                    WHEN time(`Clock-In`) > time(StartWorkTime)
                    THEN printf('%02d:%02d', 
                        (strftime('%s', time(`Clock-In`)) - strftime('%s', time(StartWorkTime))) / 3600,
                        ((strftime('%s', time(`Clock-In`)) - strftime('%s', time(StartWorkTime))) % 3600) / 60
                    )
                    ELSE '00:00'
                END AS `Late Clock In`,

                -- Early Clock In
                CASE 
                    WHEN time(`Clock-In`) < time(StartWorkTime)
                    THEN printf('%02d:%02d', 
                        (strftime('%s', time(StartWorkTime)) - strftime('%s', time(`Clock-In`))) / 3600,
                        ((strftime('%s', time(StartWorkTime)) - strftime('%s', time(`Clock-In`))) % 3600) / 60
                    )
                    ELSE '00:00'
                END AS `Early Clock In`,

                -- Early Clock Out: Use `Out` first, fallback to `Clock-Out`
                CASE 
                    WHEN (`Out` IS NOT NULL AND time(`Out`) < time(EndWorkTime))
                    THEN printf('%02d:%02d', 
                        (strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Out`))) / 3600,
                        ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Out`))) % 3600) / 60
                    )
                    WHEN (`Out` IS NULL AND `Clock-Out` IS NOT NULL AND time(`Clock-Out`) < time(EndWorkTime))
                    THEN printf('%02d:%02d', 
                        (strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Clock-Out`))) / 3600,
                        ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(`Clock-Out`))) % 3600) / 60
                    )
                    ELSE '00:00'
                END AS `Early Clock Out`,

                -- Break
                printf('%02d:%02d',
                    CASE
                        WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL AND (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`))) >= 0
                            THEN (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`)))
                        WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL
                            THEN (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`)) + 86400)
                        ELSE 0
                    END / 3600,
                    CASE
                        WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL AND (strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`))) >= 0
                            THEN ((strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`))) % 3600) / 60
                        WHEN `In` IS NOT NULL AND `Clock-Out` IS NOT NULL
                            THEN ((strftime('%s', time(`In`)) - strftime('%s', time(`Clock-Out`)) + 86400) % 3600) / 60
                        ELSE 0
                    END
                ) AS `Break`
            FROM final
        ),

        with_work_time AS (
            SELECT 
                *,
                -- Required Work Time
                printf('%02d:%02d',
                    CASE 
                        WHEN (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime))) >= 0 
                        THEN (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) - 3600)
                        ELSE (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) + 86400 - 3600)
                    END / 3600,
                    CASE 
                        WHEN (strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime))) >= 0 
                        THEN ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) - 3600) % 3600) / 60
                        ELSE ((strftime('%s', time(EndWorkTime)) - strftime('%s', time(StartWorkTime)) + 86400 - 3600) % 3600) / 60
                    END
                ) AS `Required Work Time`,

                -- Work Time
                printf('%02d:%02d',
                    CASE
                        WHEN `Out` IS NOT NULL AND (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                            THEN (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) - 3600)
                        WHEN `Out` IS NOT NULL
                            THEN (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600)
                        WHEN `Clock-Out` IS NOT NULL AND (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                            THEN (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) - 3600)
                        WHEN `Clock-Out` IS NOT NULL
                            THEN (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600)
                        ELSE 0
                    END / 3600,
                    CASE
                        WHEN `Out` IS NOT NULL AND (strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                            THEN ((strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) - 3600) % 3600) / 60
                        WHEN `Out` IS NOT NULL
                            THEN ((strftime('%s', time(`Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600) % 3600) / 60
                        WHEN `Clock-Out` IS NOT NULL AND (strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`))) >= 0
                            THEN ((strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) - 3600) % 3600) / 60
                        WHEN `Clock-Out` IS NOT NULL
                            THEN ((strftime('%s', time(`Clock-Out`)) - strftime('%s', time(`Clock-In`)) + 86400 - 3600) % 3600) / 60
                        ELSE 0
                    END
                ) AS `Work Time`
            FROM with_flags
        ),

        final_with_ot AS (
            SELECT 
                *,
                CASE 
                    WHEN `Clock-In` IS NOT NULL OR `Clock-Out` IS NOT NULL
                    THEN '00:00'
                    ELSE `Required Work Time`
                END AS `Absent`,

                CASE 
                    WHEN Workday IN ('Mon.', 'Tues.', 'Wed.', 'Thur.', 'Fri.')
                         AND (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) > 0
                    THEN printf('%02d:%02d',
                        (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) / 3600,
                        ((strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) % 3600) / 60
                    )
                    ELSE '00:00'
                END AS `OT1`,

                CASE 
                    WHEN Workday IN ('Sat.', 'Sun.')
                         AND (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) > 0
                    THEN printf('%02d:%02d',
                        (strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) / 3600,
                        ((strftime('%s', time(`Work Time`)) - strftime('%s', time(`Required Work Time`))) % 3600) / 60
                    )
                    ELSE '00:00'
                END AS `OT2`,

                '00:00' AS `OT3`
            FROM with_work_time
        )
        SELECT 
            employee_id,
            full_name,
            department,
            Date,
            Workday,
            Timetable,
            `Required Work Time`,
            StartWorkTime,
            EndWorkTime,
            `Clock-In`,
            `Clock-Out`,
            `In`,
            `Out`,
            `Late Clock In`,
            `Early Clock In`,
            `Early Clock Out`,
            `Break`,
            `Work Time`,
            `Absent`,
            `OT1`,
            `OT2`,
            `OT3`
        FROM final_with_ot
        ORDER BY employee_id, Date;
        """
        
        df = pd.read_sql_query(query, conn)
        
        if df.empty:
            raise HTTPException(status_code=404, detail="No attendance data found for the specified date range")
        
        # Get company name for title
        title_query = "SELECT cmp_name FROM hr_company LIMIT 1;"
        title_df = pd.read_sql_query(title_query, conn)
        company_name = title_df.iloc[0]['cmp_name'] if not title_df.empty else "Company Name"
        
        conn.close()
        
        # Generate Excel file using the same logic as saya.py
        output_file = generate_excel_report(df, company_name, start_date, end_date)
        
        # Generate filename based on final date range
        if start_date == end_date:
            report_filename = f"attendance_report_{start_date}.xlsx"
        else:
            report_filename = f"attendance_report_{start_date}_to_{end_date}.xlsx"
        
        return FileResponse(
            path=output_file,
            filename=report_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except sqlite3.Error as e:
        raise HTTPException(status_code=400, detail=f"Database error: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
    finally:
        # Clean up temporary database file
        if os.path.exists(temp_db_path):
            os.unlink(temp_db_path)

def generate_excel_report(df, company_name, start_date, end_date):
    """Generate Excel report using the same formatting logic as saya.py"""
    
    # Apply column swapping for NIGHT and AFTERNOON timetable rows
    # Make a copy to ensure we can modify it
    df = df.copy()
    
    # Initialize OT-F columns for all rows with default values
    df['OT1-F'] = '0.0'
    df['OT2-F'] = '0.0'
    df['OT3-F'] = '0.0'
    
    # Debug: print unique timetable values to see what we're working with
    print("DEBUG: Unique Timetable values:", df['Timetable'].unique())
    
    # ===== NIGHT SHIFT COLUMN SWAPPING =====
    # Check for NIGHT timetables (partial match to catch variations)
    night_mask_partial = df['Timetable'].str.contains('NIGHT', na=False)
    print(f"DEBUG: Found {night_mask_partial.sum()} rows containing 'NIGHT'")
    
    if night_mask_partial.any():
        print(f"DEBUG: Moving 'In' data to 'Clock-In' for {night_mask_partial.sum()} night shift rows")
        
        print("DEBUG: NIGHT - Before swap - sample row values:")
        sample_idx = df[night_mask_partial].index[0]
        print(f"  Clock-In: {df.loc[sample_idx, 'Clock-In']}")
        print(f"  In: {df.loc[sample_idx, 'In']}")
        
        # Simple swap: Move 'In' data to 'Clock-In' for night shifts only
        df.loc[night_mask_partial, 'Clock-In'] = df.loc[night_mask_partial, 'In']
        df.loc[night_mask_partial, 'In'] = None  # Clear the 'In' column for night shifts
        
        print("DEBUG: NIGHT - After swap - sample row values:")
        print(f"  Clock-In: {df.loc[sample_idx, 'Clock-In']}")
        print(f"  In: {df.loc[sample_idx, 'In']}")
        
        print("DEBUG: NIGHT column swapping completed successfully")
    else:
        print("DEBUG: No NIGHT timetables found")
    
    # ===== AFTERNOON SHIFT COLUMN SWAPPING =====
    # Check for AFTERNOON (12:00 - 00:00) timetable
    afternoon_mask = df['Timetable'] == 'AFTERNOON (12:00 - 00:00)'
    print(f"DEBUG: Found {afternoon_mask.sum()} rows with exact 'AFTERNOON (12:00 - 00:00)' match")
    
    # Also check for partial match in case of slight differences
    afternoon_mask_partial = df['Timetable'].str.contains('AFTERNOON', na=False)
    print(f"DEBUG: Found {afternoon_mask_partial.sum()} rows containing 'AFTERNOON'")
    
    # Use the exact match, but if none found, try partial
    if afternoon_mask.any():
        use_afternoon_mask = afternoon_mask
        print("DEBUG: Using exact match for AFTERNOON")
    elif afternoon_mask_partial.any():
        use_afternoon_mask = afternoon_mask_partial
        print("DEBUG: Using partial match for AFTERNOON timetables")
    else:
        use_afternoon_mask = None
        print("DEBUG: No AFTERNOON timetables found")
    
    if use_afternoon_mask is not None and use_afternoon_mask.any():
        print(f"DEBUG: Swapping columns for {use_afternoon_mask.sum()} afternoon shift rows")
        
        # Store original values for the swap
        temp_clock_out = df.loc[use_afternoon_mask, 'Clock-Out'].copy()
        temp_clock_in = df.loc[use_afternoon_mask, 'Clock-In'].copy()
        temp_in = df.loc[use_afternoon_mask, 'In'].copy()
        temp_out = df.loc[use_afternoon_mask, 'Out'].copy()
        
        print("DEBUG: AFTERNOON - Before swap - sample row values:")
        if use_afternoon_mask.any():
            sample_idx = df[use_afternoon_mask].index[0]
            print(f"  Clock-In: {df.loc[sample_idx, 'Clock-In']}")
            print(f"  Clock-Out: {df.loc[sample_idx, 'Clock-Out']}")
            print(f"  In: {df.loc[sample_idx, 'In']}")
            print(f"  Out: {df.loc[sample_idx, 'Out']}")
        
        # Perform the afternoon swap pattern:
        # Clock-Out → Clock-In
        # Clock-In → Out  
        # In → Clock-Out
        # Out → In
        df.loc[use_afternoon_mask, 'Clock-In'] = temp_clock_out    # Clock-Out → Clock-In
        df.loc[use_afternoon_mask, 'Out'] = temp_clock_in          # Clock-In → Out
        df.loc[use_afternoon_mask, 'Clock-Out'] = temp_in          # In → Clock-Out
        df.loc[use_afternoon_mask, 'In'] = temp_out                # Out → In
        
        print("DEBUG: AFTERNOON - After swap - sample row values:")
        if use_afternoon_mask.any():
            print(f"  Clock-In: {df.loc[sample_idx, 'Clock-In']}")
            print(f"  Clock-Out: {df.loc[sample_idx, 'Clock-Out']}")
            print(f"  In: {df.loc[sample_idx, 'In']}")
            print(f"  Out: {df.loc[sample_idx, 'Out']}")
        
        print("DEBUG: AFTERNOON column swapping completed successfully")
        
        # Recalculate derived columns for NIGHT shifts after swapping
        print("DEBUG: Recalculating derived columns for NIGHT shifts...")
        
        def time_to_seconds(time_str):
            """Convert HH:MM or HH:MM:SS to seconds"""
            if not time_str or pd.isna(time_str) or time_str == '':
                return None
            try:
                parts = str(time_str).split(':')
                hours = int(parts[0])
                minutes = int(parts[1])
                return hours * 3600 + minutes * 60
            except:
                return None
        
        def seconds_to_time_format(seconds):
            """Convert seconds to HH:MM format"""
            if seconds is None or seconds < 0:
                return '00:00'
            hours = seconds // 3600
            minutes = (seconds % 3600) // 60
            return f"{hours:02d}:{minutes:02d}"
        
        for idx in df[night_mask_partial].index:
            clock_in = df.loc[idx, 'Clock-In']
            clock_out = df.loc[idx, 'Clock-Out']
            in_time = df.loc[idx, 'In']
            out_time = df.loc[idx, 'Out']
            start_work_time = df.loc[idx, 'StartWorkTime']
            end_work_time = df.loc[idx, 'EndWorkTime']
            
            # Convert times to seconds for calculation
            clock_in_sec = time_to_seconds(clock_in)
            clock_out_sec = time_to_seconds(clock_out)
            in_sec = time_to_seconds(in_time)
            out_sec = time_to_seconds(out_time)
            start_sec = time_to_seconds(start_work_time)
            end_sec = time_to_seconds(end_work_time)
            
            # Recalculate Break (time between Clock-Out and In)
            if clock_out_sec is not None and in_sec is not None:
                if in_sec >= clock_out_sec:
                    break_seconds = in_sec - clock_out_sec
                else:
                    # Handle overnight case
                    break_seconds = in_sec - clock_out_sec + 86400
                df.loc[idx, 'Break'] = seconds_to_time_format(break_seconds)
            else:
                df.loc[idx, 'Break'] = '00:00'
            
            # Recalculate Late Clock In
            if clock_in_sec is not None and start_sec is not None:
                if clock_in_sec > start_sec:
                    late_seconds = clock_in_sec - start_sec
                    df.loc[idx, 'Late Clock In'] = seconds_to_time_format(late_seconds)
                else:
                    df.loc[idx, 'Late Clock In'] = '00:00'
            else:
                df.loc[idx, 'Late Clock In'] = '00:00'
            
            # Recalculate Early Clock In
            if clock_in_sec is not None and start_sec is not None:
                if clock_in_sec < start_sec:
                    early_seconds = start_sec - clock_in_sec
                    df.loc[idx, 'Early Clock In'] = seconds_to_time_format(early_seconds)
                else:
                    df.loc[idx, 'Early Clock In'] = '00:00'
            else:
                df.loc[idx, 'Early Clock In'] = '00:00'
            
            # Recalculate Early Clock Out (use Out first, fallback to Clock-Out)
            final_out_sec = out_sec if out_sec is not None else clock_out_sec
            if final_out_sec is not None and end_sec is not None:
                if final_out_sec < end_sec:
                    early_out_seconds = end_sec - final_out_sec
                    df.loc[idx, 'Early Clock Out'] = seconds_to_time_format(early_out_seconds)
                else:
                    df.loc[idx, 'Early Clock Out'] = '00:00'
            else:
                df.loc[idx, 'Early Clock Out'] = '00:00'
            
            # Recalculate Work Time (Clock-In to final out time minus 1 hour break)
            if clock_in_sec is not None and final_out_sec is not None:
                if final_out_sec >= clock_in_sec:
                    work_seconds = final_out_sec - clock_in_sec - 3600  # Subtract 1 hour break
                else:
                    # Handle overnight case
                    work_seconds = final_out_sec - clock_in_sec + 86400 - 3600
                
                if work_seconds < 0:
                    work_seconds = 0
                df.loc[idx, 'Work Time'] = seconds_to_time_format(work_seconds)
            else:
                df.loc[idx, 'Work Time'] = '00:00'
            
            # Recalculate Absent (if no clock-in or clock-out, use Required Work Time, else '00:00')
            if clock_in or clock_out:
                df.loc[idx, 'Absent'] = '00:00'
            else:
                df.loc[idx, 'Absent'] = df.loc[idx, 'Required Work Time']
            
            # Recalculate OT1, OT2, OT3 based on new Work Time
            work_time_str = df.loc[idx, 'Work Time']
            required_work_time_str = df.loc[idx, 'Required Work Time']
            workday = df.loc[idx, 'Workday']
            
            # Convert work times to seconds for comparison
            work_time_seconds = time_to_seconds(work_time_str)
            required_work_seconds = time_to_seconds(required_work_time_str)
            
            if work_time_seconds is not None and required_work_seconds is not None:
                overtime_seconds = work_time_seconds - required_work_seconds
                
                if overtime_seconds > 0:
                    overtime_str = seconds_to_time_format(overtime_seconds)
                    
                    # OT1: Overtime for weekdays (Mon-Fri)
                    if workday in ['Mon.', 'Tues.', 'Wed.', 'Thur.', 'Fri.']:
                        df.loc[idx, 'OT1'] = overtime_str
                        df.loc[idx, 'OT2'] = '00:00'
                    # OT2: Overtime for weekends (Sat-Sun)
                    elif workday in ['Sat.', 'Sun.']:
                        df.loc[idx, 'OT1'] = '00:00'
                        df.loc[idx, 'OT2'] = overtime_str
                    else:
                        df.loc[idx, 'OT1'] = '00:00'
                        df.loc[idx, 'OT2'] = '00:00'
                else:
                    # No overtime
                    df.loc[idx, 'OT1'] = '00:00'
                    df.loc[idx, 'OT2'] = '00:00'
            else:
                # Cannot calculate overtime
                df.loc[idx, 'OT1'] = '00:00'
                df.loc[idx, 'OT2'] = '00:00'
            
            # OT3 is always '00:00' according to the original logic
            df.loc[idx, 'OT3'] = '00:00'
        
        print("DEBUG: Derived columns (including OT1, OT2, OT3) recalculated successfully for NIGHT shifts")
    
    # Recalculate derived columns for AFTERNOON shifts after swapping
    if use_afternoon_mask is not None and use_afternoon_mask.any():
        print("DEBUG: Recalculating derived columns for AFTERNOON shifts...")
        
        for idx in df[use_afternoon_mask].index:
            clock_in = df.loc[idx, 'Clock-In']
            clock_out = df.loc[idx, 'Clock-Out']
            in_time = df.loc[idx, 'In']
            out_time = df.loc[idx, 'Out']
            start_work_time = df.loc[idx, 'StartWorkTime']
            end_work_time = df.loc[idx, 'EndWorkTime']
            
            # Convert times to seconds for calculation
            clock_in_sec = time_to_seconds(clock_in)
            clock_out_sec = time_to_seconds(clock_out)
            in_sec = time_to_seconds(in_time)
            out_sec = time_to_seconds(out_time)
            start_sec = time_to_seconds(start_work_time)
            end_sec = time_to_seconds(end_work_time)
            
            # Recalculate Break (time between Clock-Out and In)
            if clock_out_sec is not None and in_sec is not None:
                if in_sec >= clock_out_sec:
                    break_seconds = in_sec - clock_out_sec
                else:
                    # Handle overnight case
                    break_seconds = in_sec - clock_out_sec + 86400
                df.loc[idx, 'Break'] = seconds_to_time_format(break_seconds)
            else:
                df.loc[idx, 'Break'] = '00:00'
            
            # Recalculate Late Clock In
            if clock_in_sec is not None and start_sec is not None:
                if clock_in_sec > start_sec:
                    late_seconds = clock_in_sec - start_sec
                    df.loc[idx, 'Late Clock In'] = seconds_to_time_format(late_seconds)
                else:
                    df.loc[idx, 'Late Clock In'] = '00:00'
            else:
                df.loc[idx, 'Late Clock In'] = '00:00'
            
            # Recalculate Early Clock In
            if clock_in_sec is not None and start_sec is not None:
                if clock_in_sec < start_sec:
                    early_seconds = start_sec - clock_in_sec
                    df.loc[idx, 'Early Clock In'] = seconds_to_time_format(early_seconds)
                else:
                    df.loc[idx, 'Early Clock In'] = '00:00'
            else:
                df.loc[idx, 'Early Clock In'] = '00:00'
            
            # Recalculate Early Clock Out (use Out first, fallback to Clock-Out)
            final_out_sec = out_sec if out_sec is not None else clock_out_sec
            if final_out_sec is not None and end_sec is not None:
                if final_out_sec < end_sec:
                    early_out_seconds = end_sec - final_out_sec
                    df.loc[idx, 'Early Clock Out'] = seconds_to_time_format(early_out_seconds)
                else:
                    df.loc[idx, 'Early Clock Out'] = '00:00'
            else:
                df.loc[idx, 'Early Clock Out'] = '00:00'
            
            # Recalculate Work Time (Clock-In to final out time minus 1 hour break)
            if clock_in_sec is not None and final_out_sec is not None:
                if final_out_sec >= clock_in_sec:
                    work_seconds = final_out_sec - clock_in_sec - 3600  # Subtract 1 hour break
                else:
                    # Handle overnight case
                    work_seconds = final_out_sec - clock_in_sec + 86400 - 3600
                
                if work_seconds < 0:
                    work_seconds = 0
                df.loc[idx, 'Work Time'] = seconds_to_time_format(work_seconds)
            else:
                df.loc[idx, 'Work Time'] = '00:00'
            
            # Recalculate Absent (if no clock-in or clock-out, use Required Work Time, else '00:00')
            if clock_in or clock_out:
                df.loc[idx, 'Absent'] = '00:00'
            else:
                df.loc[idx, 'Absent'] = df.loc[idx, 'Required Work Time']
            
            # Recalculate OT1, OT2, OT3 based on new Work Time
            work_time_str = df.loc[idx, 'Work Time']
            required_work_time_str = df.loc[idx, 'Required Work Time']
            workday = df.loc[idx, 'Workday']
            
            # Convert work times to seconds for comparison
            work_time_seconds = time_to_seconds(work_time_str)
            required_work_seconds = time_to_seconds(required_work_time_str)
            
            if work_time_seconds is not None and required_work_seconds is not None:
                overtime_seconds = work_time_seconds - required_work_seconds
                
                if overtime_seconds > 0:
                    overtime_str = seconds_to_time_format(overtime_seconds)
                    
                    # OT1: Overtime for weekdays (Mon-Fri)
                    if workday in ['Mon.', 'Tues.', 'Wed.', 'Thur.', 'Fri.']:
                        df.loc[idx, 'OT1'] = overtime_str
                        df.loc[idx, 'OT2'] = '00:00'
                    # OT2: Overtime for weekends (Sat-Sun)
                    elif workday in ['Sat.', 'Sun.']:
                        df.loc[idx, 'OT1'] = '00:00'
                        df.loc[idx, 'OT2'] = overtime_str
                    else:
                        df.loc[idx, 'OT1'] = '00:00'
                        df.loc[idx, 'OT2'] = '00:00'
                else:
                    # No overtime
                    df.loc[idx, 'OT1'] = '00:00'
                    df.loc[idx, 'OT2'] = '00:00'
            else:
                # Cannot calculate overtime
                df.loc[idx, 'OT1'] = '00:00'
                df.loc[idx, 'OT2'] = '00:00'
            
            # OT3 is always '00:00' according to the original logic
            df.loc[idx, 'OT3'] = '00:00'
        
        print("DEBUG: Derived columns (including OT1, OT2, OT3) recalculated successfully for AFTERNOON shifts")
    
    # Calculate OT-F values for ALL rows (not just NIGHT shifts)
    print("DEBUG: Calculating OT-F values for all rows...")
    
    def time_to_decimal_hours_global(time_str):
        """Convert HH:MM time to decimal hours"""
        if not time_str or time_str == '00:00' or pd.isna(time_str):
            return 0.0
        try:
            parts = str(time_str).split(':')
            hours = int(parts[0])
            minutes = int(parts[1])
            return hours + (minutes / 60.0)
        except:
            return 0.0
    
    def floor_to_half_hour_global(decimal_hours):
        """Floor decimal hours to nearest 0.5 increment"""
        if decimal_hours == 0.0:
            return 0.0
        # Floor to nearest 0.5: multiply by 2, floor, then divide by 2
        return float(int(decimal_hours * 2)) / 2.0
    
    # Process all rows to calculate OT-F values
    for idx in df.index:
        # Get current OT values and convert to floored format
        ot1_decimal = time_to_decimal_hours_global(df.loc[idx, 'OT1'])
        ot2_decimal = time_to_decimal_hours_global(df.loc[idx, 'OT2'])
        ot3_decimal = time_to_decimal_hours_global(df.loc[idx, 'OT3'])
        
        # Apply floor to 0.5 hour increment
        ot1_floored = floor_to_half_hour_global(ot1_decimal)
        ot2_floored = floor_to_half_hour_global(ot2_decimal)
        ot3_floored = floor_to_half_hour_global(ot3_decimal)
        
        # Store in OT-F columns as decimal format
        df.loc[idx, 'OT1-F'] = f"{ot1_floored:.1f}" if ot1_floored > 0 else "0.0"
        df.loc[idx, 'OT2-F'] = f"{ot2_floored:.1f}" if ot2_floored > 0 else "0.0"
        df.loc[idx, 'OT3-F'] = f"{ot3_floored:.1f}" if ot3_floored > 0 else "0.0"
    
    print("DEBUG: OT-F values calculated for all rows successfully")
    
    # Add missing columns to DataFrame for proper TOTAL calculation
    print("DEBUG: Adding missing columns to DataFrame for TOTAL calculation...")
    
    for idx in df.index:
        is_sunday = df.loc[idx, 'Workday'] == 'Sun.'
        
        # Total Base: 1.0 for non-Sunday, 0.0 for Sunday
        if is_sunday:
            df.loc[idx, 'Total Base'] = '0.0'
        else:
            df.loc[idx, 'Total Base'] = '1.0'
        
        # Day: Sunday is always empty, other days depend on clock-in/out presence
        if is_sunday:
            df.loc[idx, 'Day'] = ''
        else:
            # Check if worker was present (has clock-in OR clock-out)
            clock_in = df.loc[idx, 'Clock-In']
            clock_out = df.loc[idx, 'Clock-Out']
            if clock_in or clock_out:  # If either clock-in or clock-out has value
                df.loc[idx, 'Day'] = '1.0'
            else:
                df.loc[idx, 'Day'] = ''
        
        # Total Day: Fill every cell with 1.0
        df.loc[idx, 'Total Day'] = '1.0'
        
        # Night Shift: 2.0 if NIGHT timetable, 0.0 otherwise
        original_timetable = df.loc[idx, 'Timetable']
        if 'NIGHT' in str(original_timetable).upper():
            df.loc[idx, 'Night Shift'] = '2.0'
        else:
            df.loc[idx, 'Night Shift'] = '0.0'
        
        # Penalty and Allowence: default to 0.0
        df.loc[idx, 'Penalty'] = '0.0'
        df.loc[idx, 'Allowence'] = '0.0'
        
        # Leave columns: default to empty
        for leave_col in ['H', 'MC', 'AL', 'UP', 'S']:
            df.loc[idx, leave_col] = ''
    
    print("DEBUG: Missing columns added to DataFrame successfully")
    
    # Columns to sum
    sum_columns = ['Required Work Time', 'Work Time', 'Absent', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Penalty', 'OT1', 'OT2', 'OT3', 'OT1-F', 'OT2-F', 'OT3-F', 'Night Shift', 'Allowence', 'Total Base', 'Day', 'H', 'MC', 'AL', 'UP', 'S', 'Total Day']

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Create Tahoma font
    tahoma_font = Font(name='Tahoma', size=10)
    tahoma_bold_font = Font(name='Tahoma', size=10, bold=True)

    # Create yellow fill for Sunday rows
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Create light red fill for late clock-in cells
    red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')

    # Create light purple fill for penalty cells
    light_purple_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

    # Create light orange fill for total base cells
    light_orange_fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')

    # Create bright red font for suspicious early clock in
    bright_red_font = Font(name='Tahoma', size=10, color='FF0000')

    # Create orange fill for suspicious punch pattern rows
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    # Create green fill for total rows
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    # Create light yellow fill for specific header columns
    header_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    # Create light orange fill for Total Base column
    total_base_fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')

    # Create light green fill for Total Day column
    total_day_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    # Create light orange fill for Clock columns
    clock_header_fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')

    # Create light blue fill for OT columns
    ot_header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

    # Create light purple fill for Penalty column
    penalty_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')

    # Create title font
    title_font = Font(name='Tahoma', size=22, bold=True)

    # Create subtitle font
    subtitle_font = Font(name='Tahoma', size=14, bold=True)

    current_row = 1

    # Add company name title at the top
    ws.cell(row=current_row, column=1, value=company_name).font = title_font
    current_row += 2  # Add 1 empty row of space after title

    # Add subtitle with date range
    ws.cell(row=current_row, column=1, value=f"Monthly Statement Report ({start_date} to {end_date})").font = subtitle_font
    current_row += 1
    
    # Add note about expandable columns
    note_font = Font(name='Tahoma', size=9, italic=True)
    ws.cell(row=current_row, column=1, value="📍 Note: OT columns (OT1,OT2,OT3) and Leave columns (H,MC,AL,UP,S) are grouped. Click [+] buttons to expand.").font = note_font
    current_row += 2  # Add 1 empty row of space after note

    # Write column headers only once
    cols_to_show = ['Date', 'Workday', 'Timetable', 'EYEE NAME', 'StartWorkTime', 'EndWorkTime', 'Clock-In', 'Clock-Out', 'In', 'Out', 'Required Work Time', 'Break', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Work Time', 'Absent', 'Penalty', 'OT1', 'OT2', 'OT3', 'OT1-F', 'OT2-F', 'OT3-F', 'Night Shift', 'Allowence', 'Total Base', 'Day', 'H', 'MC', 'AL', 'UP', 'S', 'Total Day']
    # Define columns that need the special header color
    colored_header_columns = ['Date', 'Workday', 'Timetable', 'EYEE NAME', 'StartWorkTime', 'EndWorkTime', 'Day', 'H', 'MC', 'AL', 'UP', 'S', 'Required Work Time', 'Break', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Work Time', 'Absent']
    for col_idx, col_name in enumerate(cols_to_show, start=1):
        header_cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        header_cell.font = tahoma_bold_font
        header_cell.alignment = Alignment(wrap_text=True)
        # Apply background color to specific columns
        if col_name in colored_header_columns:
            header_cell.fill = header_fill
        elif col_name == 'Total Base':
            header_cell.fill = total_base_fill
        elif col_name in ['Total Day', 'Night Shift', 'Allowence']:
            header_cell.fill = total_day_fill
        elif col_name in ['Clock-In', 'Clock-Out', 'In', 'Out']:
            header_cell.fill = clock_header_fill
        elif col_name in ['OT1', 'OT2', 'OT3', 'OT1-F', 'OT2-F', 'OT3-F']:
            header_cell.fill = ot_header_fill
        elif col_name == 'Penalty':
            header_cell.fill = penalty_fill
    ws.row_dimensions[current_row].height = 39.75
    current_row += 1

    # Function to format time from HH:MM:SS to HH:MM
    def format_time(time_str):
        if not time_str or time_str == '':
            return time_str
        try:
            # If time has seconds, remove them
            if time_str.count(':') == 2:
                parts = time_str.split(':')
                return f"{parts[0]}:{parts[1]}"
            else:
                return time_str
        except:
            return time_str

    for emp_id, group in df.groupby('employee_id'):
        emp_info = group.iloc[0]

        # Write employee info in single row format
        ws.cell(row=current_row, column=1, value="Employee ID").font = tahoma_bold_font
        ws.cell(row=current_row, column=2, value=emp_id).font = tahoma_font
        ws.cell(row=current_row, column=3, value="Full Name").font = tahoma_bold_font
        ws.cell(row=current_row, column=4, value=emp_info['full_name']).font = tahoma_font
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Write data rows
        data_start_row = current_row  # Mark the start of data rows for grouping
        for _, row in group.iterrows():
            is_sunday = row.get('Workday') == 'Sun.'
            
            # Function to convert time string to minutes for comparison
            def time_to_minutes(time_str):
                if not time_str or time_str in ['', '0:00', '00:00']:
                    return 0
                try:
                    parts = time_str.split(':')
                    hours = int(parts[0])
                    minutes = int(parts[1])
                    return hours * 60 + minutes
                except:
                    return 0
            
            # Check if this is a suspicious row (Early Clock In > 2:30)
            early_in_value = row.get('Early Clock In', "")
            early_in_minutes = time_to_minutes(early_in_value)
            is_suspicious_row = early_in_minutes > 150  # 2:30 = 150 minutes
            
            # Check for suspicious punch patterns
            def has_value(val):
                return val and val.strip() != ''
            
            clock_in = row.get('Clock-In', '')
            clock_out = row.get('Clock-Out', '')
            in_punch = row.get('In', '')
            out_punch = row.get('Out', '')
            
            clock_in_exists = has_value(clock_in)
            clock_out_exists = has_value(clock_out)
            in_exists = has_value(in_punch)
            out_exists = has_value(out_punch)
            
            # Determine if punch pattern is suspicious
            is_punch_suspicious = False
            if clock_in_exists and not clock_out_exists and not in_exists and not out_exists:
                # Only Clock In exists = SUSPICIOUS
                is_punch_suspicious = True
            elif clock_in_exists and clock_out_exists and in_exists and not out_exists:
                # Clock In + Clock Out + In (missing Out) = SUSPICIOUS
                is_punch_suspicious = True
            
            for col_idx, col_name in enumerate(cols_to_show, start=1):
                # Set value based on column name
                if col_name == 'EYEE NAME':
                    cell_value = emp_info['full_name']
                elif col_name == 'Timetable':
                    # Format as "Timetable (StartTime - EndTime)"
                    timetable_name = row.get('Timetable', "")
                    start_time = format_time(row.get('StartWorkTime', ""))
                    end_time = format_time(row.get('EndWorkTime', ""))
                    if timetable_name and start_time and end_time:
                        cell_value = f"{timetable_name} ({start_time} - {end_time})"
                    else:
                        cell_value = timetable_name
                elif col_name == 'Penalty':
                    cell_value = '0.0'
                elif col_name == 'Night Shift':
                    # Check if original Timetable contains "NIGHT"
                    original_timetable = row.get('Timetable', "")
                    if 'NIGHT' in str(original_timetable).upper():
                        cell_value = '2.0'
                    else:
                        cell_value = '0.0'
                elif col_name == 'Allowence':
                    cell_value = '0.0'
                elif col_name == 'Total Base':
                    # Set 1.0 for non-Sunday, 0.0 for Sunday
                    if is_sunday:
                        cell_value = '0.0'
                    else:
                        cell_value = '1.0'
                elif col_name == 'Day':
                    # Sunday is always empty, other days depend on clock-in/out presence
                    if is_sunday:
                        cell_value = ''
                    else:
                        # Check if worker was present (has clock-in OR clock-out)
                        clock_in = row.get('Clock-In', '')
                        clock_out = row.get('Clock-Out', '')
                        if clock_in or clock_out:  # If either clock-in or clock-out has value
                            cell_value = '1.0'
                        else:
                            cell_value = ''
                elif col_name in ['H', 'MC', 'AL', 'UP', 'S']:
                    # Leave these columns empty
                    cell_value = ''
                elif col_name == 'Total Day':
                    # Fill every cell with 1.0
                    cell_value = '1.0'
                elif col_name in ['StartWorkTime', 'EndWorkTime', 'Clock-In', 'Clock-Out', 'In', 'Out']:
                    # Format time columns to show only HH:MM (remove seconds)
                    cell_value = format_time(row.get(col_name, ""))
                else:
                    cell_value = row.get(col_name, "")
                
                # Convert OT time values to decimal format
                if col_name in ['OT1', 'OT2', 'OT3'] and cell_value:
                    # Convert "hh:mm" to decimal (e.g., "02:30" -> "2.5")
                    def time_to_decimal(time_str):
                        if not time_str or time_str in ['', '0:00', '00:00']:
                            return '0.0'
                        try:
                            parts = time_str.split(':')
                            hours = int(parts[0])
                            minutes = int(parts[1])
                            decimal_value = hours + (minutes / 60.0)
                            return f"{decimal_value:.2f}".rstrip('0').rstrip('.')
                        except:
                            return cell_value
                    
                    cell_value = time_to_decimal(cell_value)
                
                # Handle OT-F columns (floored overtime values)
                if col_name in ['OT1-F', 'OT2-F', 'OT3-F']:
                    # For OT-F columns, get value from DataFrame, handle NaN and None cases
                    val = row.get(col_name, '0.0')
                    if pd.isna(val) or val is None or val == '' or str(val).lower() == 'nan':
                        cell_value = '0.0'
                    else:
                        cell_value = str(val)
                
                cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                
                # Apply font styling based on suspicious row detection
                if is_suspicious_row and col_name in ['Timetable', 'StartWorkTime', 'EndWorkTime', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Night Shift']:
                    cell.font = bright_red_font  # Apply bright red font for suspicious rows
                else:
                    cell.font = tahoma_font  # Apply regular font to data cells
                
                # Apply cell coloring based on conditions (punch suspicious takes priority)
                if is_punch_suspicious:
                    # Apply orange background to entire row for suspicious punch patterns
                    cell.fill = orange_fill
                elif col_name == 'Late Clock In':
                    # Check if this is a late clock-in cell with a value other than 0:00
                    late_value = row.get(col_name, "")
                    if late_value and late_value not in ['0:00', '00:00', '']:
                        cell.fill = red_fill
                    elif is_sunday:
                        cell.fill = yellow_fill
                elif col_name == 'Early Clock In':
                    # Apply Sunday background if applicable
                    if is_sunday:
                        cell.fill = yellow_fill
                elif col_name == 'Early Clock Out':
                    # Check if this is an early clock-out cell with a value other than 0:00
                    early_out_value = row.get(col_name, "")
                    if early_out_value and early_out_value not in ['0:00', '00:00', '']:
                        cell.fill = red_fill  # Light red even on Sunday
                    elif is_sunday:
                        cell.fill = yellow_fill
                elif col_name == 'Penalty':
                    # Check if this is a penalty cell
                    if is_sunday:
                        cell.fill = yellow_fill
                    elif cell_value and cell_value != '':
                        # Apply light purple color to all non-empty, non-Sunday cells
                        cell.fill = penalty_fill
                elif col_name == 'Total Base':
                    # Check if this is a total base cell
                    if is_sunday:
                        cell.fill = yellow_fill
                    elif cell_value and cell_value != '':
                        # Apply new orange color to all non-empty, non-Sunday cells
                        cell.fill = total_base_fill
                elif col_name == 'Total Day':
                    # Check if this is a total day cell
                    if is_sunday:
                        cell.fill = yellow_fill
                    elif cell_value and cell_value != '':
                        # Apply light green color to all non-empty, non-Sunday cells
                        cell.fill = total_day_fill
                elif col_name in ['Night Shift', 'Allowence']:
                    # Check if this is a night shift or allowence cell
                    if is_sunday:
                        cell.fill = yellow_fill
                    elif cell_value and cell_value != '':
                        # Apply light green color to all non-empty, non-Sunday cells
                        cell.fill = total_day_fill
                elif is_sunday:
                    cell.fill = yellow_fill
            ws.row_dimensions[current_row].height = 18
            # Set outline level for data rows (level 1 for grouping)
            ws.row_dimensions[current_row].outline_level = 1
            current_row += 1

        # Add TOTAL row for each employee
        def time_to_minutes_for_sum(time_str):
            """Convert HH:MM time string to minutes for summing"""
            if not time_str or time_str in ['', '0:00', '00:00']:
                return 0
            try:
                parts = str(time_str).split(':')
                hours = int(parts[0])
                minutes = int(parts[1])
                return hours * 60 + minutes
            except:
                return 0
        
        def minutes_to_time_str(total_minutes):
            """Convert total minutes back to HH:MM format"""
            if total_minutes == 0:
                return '00:00'
            hours = total_minutes // 60
            minutes = total_minutes % 60
            return f"{hours:02d}:{minutes:02d}"
        
        def sum_decimal_values(group, col_name):
            """Sum decimal values from a column"""
            total = 0.0
            has_values = False
            
            for _, row in group.iterrows():
                val = row.get(col_name, '')
                
                if val and str(val).strip() not in ['', '0.0', 'nan', 'None', 'NaN']:
                    try:
                        # Handle different data formats
                        val_str = str(val).strip()
                        
                        # Special handling for certain column types
                        if col_name in ['OT1', 'OT2', 'OT3']:
                            # These might be in decimal format already (e.g., "2.5") or time format (e.g., "02:30")
                            if ':' in val_str:
                                # Convert time format to decimal hours
                                parts = val_str.split(':')
                                hours = int(parts[0])
                                minutes = int(parts[1])
                                numeric_val = hours + (minutes / 60.0)
                            else:
                                numeric_val = float(val_str)
                        else:
                            # Standard decimal conversion
                            numeric_val = float(val_str)
                        
                        total += numeric_val
                        if numeric_val != 0.0:
                            has_values = True
                            
                    except (ValueError, TypeError):
                        pass
            
            # Return the total if there are any values, otherwise return "0.0" for columns that should show totals
            if has_values or total > 0:
                return f"{total:.1f}"
            else:
                # For certain columns, always show 0.0 even if no values
                always_show_zero = ['Penalty', 'OT1', 'OT2', 'OT3', 'OT1-F', 'OT2-F', 'OT3-F', 'Night Shift', 'Allowence', 'Total Base', 'Day', 'Total Day']
                return "0.0" if col_name in always_show_zero else ""
        
        # Calculate totals for time-based columns
        time_columns = ['Required Work Time', 'Work Time', 'Absent', 'Late Clock In', 'Early Clock In', 'Early Clock Out', 'Break']
        decimal_columns = ['Penalty', 'OT1', 'OT2', 'OT3', 'OT1-F', 'OT2-F', 'OT3-F', 'Night Shift', 'Allowence', 'Total Base', 'Day', 'H', 'MC', 'AL', 'UP', 'S', 'Total Day']
        
        # Columns that should show count of non-empty records
        count_columns = ['Workday', 'Timetable']
        
        # Columns that should remain empty in totals
        empty_columns = ['EYEE NAME', 'StartWorkTime', 'EndWorkTime', 'Clock-In', 'Clock-Out', 'In', 'Out']
        
        # Write TOTAL row
        for col_idx, col_name in enumerate(cols_to_show, start=1):
            if col_name == 'Date':
                cell_value = 'TOTAL'
            elif col_name in time_columns:
                # Sum time values
                total_minutes = 0
                for _, row in group.iterrows():
                    total_minutes += time_to_minutes_for_sum(row.get(col_name, ''))
                cell_value = minutes_to_time_str(total_minutes)
            elif col_name in decimal_columns:
                # Sum decimal values
                cell_value = sum_decimal_values(group, col_name)
            elif col_name in count_columns:
                # Count non-empty values
                count = 0
                for _, row in group.iterrows():
                    val = row.get(col_name, '')
                    if val and str(val).strip() != '':
                        count += 1
                cell_value = str(count) if count > 0 else ''
            elif col_name in empty_columns:
                # Leave these columns empty in total row
                cell_value = ''
            else:
                # For any remaining columns, try to sum if they contain numeric/time data
                # First try as decimal
                try:
                    total_decimal = 0.0
                    has_values = False
                    for _, row in group.iterrows():
                        val = row.get(col_name, '')
                        if val and str(val).strip() not in ['', '0.0', 'nan']:
                            try:
                                total_decimal += float(val)
                                has_values = True
                            except:
                                pass
                    if has_values:
                        cell_value = f"{total_decimal:.1f}" if total_decimal > 0 else "0.0"
                    else:
                        # Try as time format
                        total_minutes = 0
                        has_time_values = False
                        for _, row in group.iterrows():
                            val = row.get(col_name, '')
                            if val and str(val).strip() not in ['', '00:00', '0:00']:
                                minutes = time_to_minutes_for_sum(val)
                                if minutes > 0:
                                    total_minutes += minutes
                                    has_time_values = True
                        cell_value = minutes_to_time_str(total_minutes) if has_time_values else ''
                except:
                    cell_value = ''
            
            cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
            cell.font = tahoma_bold_font
            cell.fill = green_fill  # Green background for total rows
            cell.border = thin_border
        
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Add one empty row after each employee
        current_row += 1

    # Apply border to all cells with data (excluding title rows)
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(cols_to_show)):
        for cell in row:
            cell.border = thin_border

    # Freeze panes to keep header rows visible when scrolling
    # Freeze at the row right after the column headers (row 6)
    ws.freeze_panes = 'A6'

    # Add column grouping for expandable columns
    # Find the column indices for OT1, OT2, OT3
    ot1_col_idx = None
    ot2_col_idx = None
    ot3_col_idx = None
    
    # Find the column indices for H, MC, AL, UP, S
    h_col_idx = None
    mc_col_idx = None
    al_col_idx = None
    up_col_idx = None
    s_col_idx = None
    
    for col_idx, col_name in enumerate(cols_to_show, start=1):
        if col_name == 'OT1':
            ot1_col_idx = col_idx
        elif col_name == 'OT2':
            ot2_col_idx = col_idx
        elif col_name == 'OT3':
            ot3_col_idx = col_idx
        elif col_name == 'H':
            h_col_idx = col_idx
        elif col_name == 'MC':
            mc_col_idx = col_idx
        elif col_name == 'AL':
            al_col_idx = col_idx
        elif col_name == 'UP':
            up_col_idx = col_idx
        elif col_name == 'S':
            s_col_idx = col_idx
    
    # Group OT1, OT2, OT3 columns together if all found
    if ot1_col_idx and ot2_col_idx and ot3_col_idx:
        start_col = min(ot1_col_idx, ot2_col_idx, ot3_col_idx)
        end_col = max(ot1_col_idx, ot2_col_idx, ot3_col_idx)
        
        # Set outline level for the OT columns (level 1 for grouping)
        for col_num in range(start_col, end_col + 1):
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[col_letter].outline_level = 1
            # Set columns to be hidden by default (collapsed state)
            ws.column_dimensions[col_letter].hidden = True
        
        print(f"DEBUG: Grouped OT columns {start_col} to {end_col} for expand/collapse functionality")
    
    # Group H, MC, AL, UP, S columns together if all found
    leave_columns = [h_col_idx, mc_col_idx, al_col_idx, up_col_idx, s_col_idx]
    if all(col is not None for col in leave_columns):
        start_col = min(leave_columns)
        end_col = max(leave_columns)
        
        # Set outline level for the leave columns (level 1 for grouping)
        for col_num in range(start_col, end_col + 1):
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[col_letter].outline_level = 1
            # Set columns to be hidden by default (collapsed state)
            ws.column_dimensions[col_letter].hidden = True
        
        print(f"DEBUG: Grouped Leave columns (H,MC,AL,UP,S) {start_col} to {end_col} for expand/collapse functionality")

    # Set outline groups to be collapsed by default
    # Hide all data rows (outline level 1) by default to show collapsed view
    for row_num in range(1, ws.max_row + 1):
        if row_num in ws.row_dimensions:
            if hasattr(ws.row_dimensions[row_num], 'outline_level') and ws.row_dimensions[row_num].outline_level == 1:
                ws.row_dimensions[row_num].hidden = True

    # Save Excel file with random number
    random_num = random.randint(1000, 9999)
    filename = f"attendance_report_{random_num}.xlsx"
    wb.save(filename)
    
    return filename

@app.get("/")
async def root():
    """Root endpoint with API information"""
    return {
        "message": "Attendance Report Generator API",
        "version": "1.0.0",
        "endpoints": {
            "POST /generate-attendance-report": "Generate Excel attendance report from ZK.db file"
        }
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 